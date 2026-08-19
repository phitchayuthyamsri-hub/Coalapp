# -*- coding: utf-8 -*-
"""Read a subcontractor's daily readiness sheet.

Tolerant on purpose: subcontractors send their own layouts. Rather than demand
the exact template, find the header row by looking for a Plate column, then map
whatever columns are present. A sheet that predates the template still imports -
it just carries less information.
"""
import re
from datetime import datetime, time

from openpyxl import load_workbook

# header text -> canonical field. Lowercased, punctuation stripped, before match.
HEADER_MAP = {
    "no": "no", "stt": "no",
    "plate": "plate", "licenseplate": "plate", "bienso": "plate", "truck": "plate",
    "location": "location", "vitri": "location",
    "status": "status", "trangthai": "status",
    "activity": "activity", "note": "activity", "ghichu": "activity",
    "timearrivemine": "arrive_time", "timearrivalmine": "arrive_time",
    "arrivetime": "arrive_time", "giodenmo": "arrive_time",
    # older BBC sheets head these columns "Arrive Mine" / "Entry Mine Date"
    "arrivemine": "arrive_time", "arrivalmine": "arrive_time",
    "datearrivemine": "arrive_date", "datearrivalmine": "arrive_date",
    "arrivedate": "arrive_date", "ngaydenmo": "arrive_date",
    "entryminedate": "arrive_date", "entrymine": "arrive_date",
    "backinservice": "back_in_service", "remark": "remark", "remarks": "remark",
}

PLATE_RE = re.compile(r"^\s*\d{2}\s*[A-Za-z]", re.I)


def _norm_header(v):
    return re.sub(r"[^a-z0-9]", "", str(v or "").lower())


def norm_plate(v):
    return re.sub(r"[^A-Za-z0-9]", "", str(v or "")).upper()


def _as_hhmm(v):
    if v is None or v == "":
        return ""
    if isinstance(v, time):
        return "%02d:%02d" % (v.hour, v.minute)
    if isinstance(v, datetime):
        return "%02d:%02d" % (v.hour, v.minute)
    m = re.match(r"^\s*(\d{1,2})[:h.](\d{2})", str(v))
    return "%02d:%s" % (int(m.group(1)), m.group(2)) if m else ""


def _as_date(v):
    if v is None or v == "":
        return ""
    if isinstance(v, datetime):
        return v.strftime("%Y-%m-%d")
    s = str(v).strip()
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y", "%d/%m", "%d-%m"):
        try:
            d = datetime.strptime(s, fmt)
            return d.strftime("%Y-%m-%d") if d.year > 1900 else ""
        except ValueError:
            pass
    return ""


def _find_header(ws, limit=30):
    """The header row is the first one containing a recognisable Plate column."""
    for r in range(1, min(ws.max_row, limit) + 1):
        cols = {}
        for c in range(1, min(ws.max_column, 20) + 1):
            f = HEADER_MAP.get(_norm_header(ws.cell(row=r, column=c).value))
            if f:
                cols[f] = c
        if "plate" in cols:
            return r, cols
    return None, {}


def parse(path):
    """-> {'rows': [...], 'header_row': n, 'columns': {...}, 'warnings': [...]}"""
    wb = load_workbook(path, data_only=True)
    ws = None
    for name in ("Readiness", "Sheet1"):
        if name in wb.sheetnames:
            ws = wb[name]
            break
    ws = ws or wb.worksheets[0]

    hrow, cols = _find_header(ws)
    warnings = []
    if not hrow:
        return {"rows": [], "header_row": None, "columns": {}, "sheet": ws.title,
                "warnings": ["No column headed 'Plate' was found - is this the "
                             "right sheet?"]}

    for wanted, label in (("arrive_time", "Time arrive Mine"),
                          ("activity", "Activity / Note"),
                          ("status", "Status")):
        if wanted not in cols:
            warnings.append("No '%s' column - that information will be blank." % label)

    def cell(r, field):
        c = cols.get(field)
        return ws.cell(row=r, column=c).value if c else None

    rows, seen = [], {}
    for r in range(hrow + 1, ws.max_row + 1):
        raw = cell(r, "plate")
        if raw is None or not PLATE_RE.match(str(raw)):
            continue
        key = norm_plate(raw)
        if not key:
            continue
        if key in seen:
            warnings.append("%s appears more than once - the later row was used."
                            % str(raw).strip())
        rec = {
            "plate": str(raw).strip(),
            "key": key,
            "location": str(cell(r, "location") or "").strip(),
            "status": str(cell(r, "status") or "").strip(),
            "activity": str(cell(r, "activity") or "").strip(),
            "arrive_time": _as_hhmm(cell(r, "arrive_time")),
            "arrive_date": _as_date(cell(r, "arrive_date")),
            "back_in_service": _as_date(cell(r, "back_in_service")),
            "remark": str(cell(r, "remark") or "").strip(),
            "row": r,
        }
        seen[key] = rec
        rows.append(rec)

    # later row wins on duplicates
    rows = list({r["key"]: r for r in rows}.values())
    return {"rows": rows, "header_row": hrow, "columns": sorted(cols),
            "sheet": ws.title, "warnings": warnings}


# Words that mean the truck is NOT available to run today. Matched as substrings,
# because subcontractors write sentences, not values: a real sheet carried
# "Unloading -> Maintenace -> 06/06 back to Mine", which an exact match treated as
# a working truck. The misspelling is deliberate - it is what they actually send.
NOT_RUNNING = ("maintenance", "maintenace", "maintainance", "not available",
               "standby", "stand by", "breakdown", "break down", "repair",
               "workshop", "garage")


def is_running(activity):
    a = str(activity or "").strip().lower()
    if not a:
        return True
    return not any(w in a for w in NOT_RUNNING)
