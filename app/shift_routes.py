"""Shift Board — answers each shift's agenda from geofence visits.

A check is not a new kind of detection: recompute_sequences already timestamps
every corridor event, so a check is a lookup on one sequence field. Shifts and
their checks are rows in the database, so NT can retime or drop a shift without
a code change.
"""
import re
from datetime import datetime, timedelta

from flask import Blueprint, jsonify, request, Response
from flask_login import login_required, current_user

from . import engine
from .models import (db, Anchor, GpsPing, Truck, Shift, ShiftCheck,
                     DailyList, DailyListRow, Subcontractor, FleetCommitment,
                     PlanSetting, RouteLeg, PlanSnapshot)
from . import readiness_import
from . import planner

bp = Blueprint("shift", __name__, url_prefix="/api/shift")

# The page itself sits outside the /api prefix, same as /gps-capture.
page_bp = Blueprint("shift_page", __name__)


def _no_store(resp):
    """Every response here depends on WHO is asking. A cached copy served after
    someone logs in as a different user shows them the previous user's role and
    permissions, so none of it may be stored."""
    resp.headers["Cache-Control"] = "no-store, no-cache, must-revalidate, max-age=0"
    resp.headers["Pragma"] = "no-cache"
    resp.headers["Expires"] = "0"
    return resp


bp.after_request(_no_store)
page_bp.after_request(_no_store)


@page_bp.route("/shift-board")
@login_required
def shift_board_page():
    from flask import render_template
    return render_template("shift_board.html")


@page_bp.route("/planner")
@login_required
def planner_page():
    from flask import render_template
    return render_template("dispatch_planner.html")


@page_bp.route("/ops")
@login_required
def ops_page():
    """Readiness, the plan and the monitor on one page.

    They are one job seen from three sides, and sending people to three URLs to
    do it made the day harder than it is. Each view is kept exactly as it is and
    embedded here, so nothing about how they work changes - only where they sit.
    """
    from flask import render_template
    return render_template("ops.html")


@page_bp.route("/monitor")
@login_required
def monitor_page():
    from flask import render_template
    return render_template("monitor.html")


@page_bp.route("/readiness")
@login_required
def readiness_page():
    """The readiness page's own name. /shift-board still answers, because it is
    bookmarked and linked to, but the monitoring name belongs to /monitor."""
    from flask import render_template
    return render_template("shift_board.html")


# check code -> (anchor role, which edge of the visit, human phrase)
# Evidence comes from visits directly, not from cycles: a cycle used to require a
# mine visit to exist at all, which meant a truck that never arrived produced no
# row and no alarm. The plan opens the cycle now; GPS only confirms it.
CHECK_SPEC = {
    "arrive_mine":   ("xppl",    "enter", "arrived at XPPL Mine"),
    "load_done":     ("loading", "exit",  "finished loading"),
    "depart_border": ("border",  "exit",  "departed Lalay border"),
    "arrive_port":   ("port",    "enter", "arrived at Chan May Port"),
    "unload_done":   ("port",    "exit",  "finished unloading at port"),
    "depart_ql49":   ("ql49",    "exit",  "departed QL49"),
}

CYCLE_SPAN = timedelta(days=3)   # a loop runs 48h, 72h if the QL49 window is used

LOCAL_OFFSET = timedelta(hours=7)   # times are shown UTC+7, as elsewhere


def _role():
    return (getattr(current_user, "role", "") or "monitor").lower()


def _can(action, state):
    """Who may do what, given the list's current state.

    A confirmed list is locked: only a manager or an admin may reopen or edit it.
    A supervisor may build and submit, but may not approve their own submission.
    """
    r = _role()
    # An admin stands in for any role, but the state machine still applies:
    # confirming a list that was never submitted would leave no record of who
    # sent it, which is the whole point of the chain.
    adm = (r == "admin")
    if action == "edit":
        # Rows keep their own state, so a supervisor can go on working the
        # pending ones even after part of the list has been approved.
        return adm or r in ("supervisor", "manager")
    if action == "submit":
        # Always available: more trucks can be sent after an earlier batch.
        return (adm or r == "supervisor") and state != "none"
    if action in ("confirm", "reject"):
        return (adm or r == "manager") and state == "submitted"
    if action == "edit_time":
        # Only the supervisor sets arrival timing. The manager reviews and
        # confirms; letting both edit it means nobody owns the number.
        return (adm or r == "supervisor") and state in ("draft", "rejected", "none")
    if action == "reopen":
        # Withdrawn by design: a list is confirmed or it is not. Changes after
        # confirmation go through the per-row pending/reject flow instead.
        return False
    return False


def _req_sub_id(d=None):
    """Which company's list is being asked for.

    A subcontractor login is pinned to its own company whatever the request
    says - the caller never chooses. Everyone else picks, and None means the
    legacy pre-per-company list.
    """
    if _role() == "subcontractor":
        return getattr(current_user, "subcontractor_id", None)
    if d is not None and "subcontractor_id" in d:
        v = d.get("subcontractor_id")
        return int(v) if str(v or "").strip().isdigit() else None
    v = request.args.get("subcontractor_id")
    return int(v) if str(v or "").strip().isdigit() else None


def _asked_all(d=None):
    """True when the caller asked for every company at once.

    Never for a subcontractor login: that account is pinned to its own company
    and must not be handed the whole corridor's day. Kept separate from
    _req_sub_id because a missing id already means the legacy pre-per-company
    list - if "all" fell through to that, a write would edit the wrong list
    instead of being refused.
    """
    if _role() == "subcontractor":
        return False
    if d is not None and "subcontractor_id" in d:
        v = d.get("subcontractor_id")
    else:
        v = request.args.get("subcontractor_id")
    return str(v or "").strip().lower() == "all"


def _one_company_only():
    return jsonify(error="Pick one company. A list is uploaded, submitted and "
                         "confirmed for a single company, so there is no list "
                         "for “all” to act on."), 400


def _find_list(day, sub_id):
    return DailyList.query.filter_by(list_date=day, subcontractor_id=sub_id).first()


def _all_payload(day):
    """Every company's list for one day, merged into one read-only view.

    The day is worked one company at a time, but it is READ whole: who is going
    tomorrow is a question about the corridor, not about Bao Binh. So this
    merges the rows and reports each company's own state alongside them.

    `can` is false throughout on purpose. Upload, Submit and Confirm each act on
    one company's list, and a button here would have to invent which list it
    meant - so the view carries none of them and says so instead.
    """
    subs = {s.id: (s.short or s.name) for s in Subcontractor.query.all()}
    lists = DailyList.query.filter_by(list_date=day).all()
    rows, parties, applied = [], [], 0
    for dl in sorted(lists, key=lambda d: subs.get(d.subcontractor_id, "")):
        p = _list_payload(dl, day)
        rows.extend(p["rows"])
        applied += p["applied_count"]
        # Trucks ON the sheet: the absent ones ride along as rows so they can be
        # seen, but counting them here would say a company sent more than it did.
        parties.append({"id": dl.subcontractor_id,
                        "short": subs.get(dl.subcontractor_id, "(no company)"),
                        "state": dl.state,
                        "trucks": sum(1 for r in p["rows"] if not r.get("absent"))})
    # Merged rows sort the way a single list does: by when the truck is due at
    # the mine, with the untimed ones - the trucks not going - at the bottom.
    rows.sort(key=lambda r: (0 if r["arrive_date"] else 1, r["arrive_date"],
                             0 if r["arrive"] else 1, r["arrive"], r["plate"]))
    states = {p["state"] for p in parties}
    state = "none" if not states else (states.pop() if len(states) == 1 else "mixed")
    return {
        "date": day, "all": True, "state": state, "rows": rows,
        "parties": parties, "applied_count": applied,
        "submitted_by": None, "submitted_at": None,
        "confirmed_by": None, "confirmed_at": None, "reject_reason": "",
        "role": _role(),
        "can": {a: False for a in ("edit", "submit", "confirm", "reject",
                                   "edit_time", "reopen")},
    }


def _restate(dl):
    """The list's state follows its rows, because trucks move in batches.

    A supervisor can send five trucks now and eight more an hour later, so
    "submitted" is not a property of the whole list - it means "some rows are
    sitting with the manager". Derived, never set by hand.
    """
    rows = DailyListRow.query.filter_by(list_id=dl.id).all()
    st = {r.state or "pending" for r in rows}
    if "applied" in st:
        dl.state = "submitted"          # something is waiting on the manager
    elif "approved" in st or "denied" in st:
        dl.state = "confirmed"          # decided, nothing outstanding
    else:
        dl.state = "draft"
    return dl.state


def _local(dt):
    """Stored UTC, shown UTC+7. Every time a person reads goes through here."""
    return (dt + LOCAL_OFFSET) if dt else None


def _fmt(dt):
    return _local(dt).strftime("%Y-%m-%d %H:%M") if dt else None


def _day_bounds(day):
    """The UTC window covering one LOCAL day. Raises ValueError on a bad date."""
    lo = datetime.strptime(day, "%Y-%m-%d") - LOCAL_OFFSET
    return lo, lo + timedelta(days=1)


def _shift_deadline(day, shift):
    """When the shift is over, in UTC.

    A night shift that ends before it starts rolls over to the next day.
    """
    start = datetime.strptime(day + " " + (shift.start_hhmm or "00:00"),
                              "%Y-%m-%d %H:%M")
    end = datetime.strptime(day + " " + (shift.end_hhmm or "23:59"),
                            "%Y-%m-%d %H:%M")
    if end <= start:
        end += timedelta(days=1)
    return end - LOCAL_OFFSET


def _visits_and_roles():
    """Every geofence visit the GPS supports, and which anchor plays which role.

    Built the same way /api/visits builds them, so a check answered here and a
    visit listed there can never disagree.
    """
    anchors = [{"id": a.id, "name": a.name, "polygon": a.polygon,
                "min_dwell_min": a.min_dwell_min} for a in Anchor.query.all()]
    roles = {a.role: a.id for a in Anchor.query.all() if a.role}
    pings = [{"plate": p.plate, "dt": p.dt, "lat": p.lat, "lng": p.lng,
              "speed": p.speed, "status": p.status}
             for p in GpsPing.query.order_by(GpsPing.dt).all()]
    deactivated = {engine.norm_plate(t.plate) for t in
                   Truck.query.filter_by(status="deactivated").all()}
    return engine.build_visits(pings, anchors, deactivated), roles


def _list_payload(dl, day):
    state = dl.state if dl else "none"
    rows = []
    if dl:
        recs = DailyListRow.query.filter_by(list_id=dl.id).all()
        # Sort by arrival date then time, A-Z. Rows with no timing sort last -
        # they are the ones not running, and they belong at the bottom.
        def _k(r):
            d, t = (r.arrive_date or ""), (r.arrive_hhmm or "")
            return (0 if d else 1, d, 0 if t else 1, t, r.plate or "")
        recs.sort(key=_k)
        s = db.session.get(Subcontractor, dl.subcontractor_id) \
            if dl.subcontractor_id else None
        short = (s.short or s.name) if s else ""
        roster = {}
        if dl.subcontractor_id:
            roster = {c.key: c.plate for c in FleetCommitment.query.filter_by(
                subcontractor_id=dl.subcontractor_id, released_on="").all()}
        for r in recs:
            rows.append({
                "plate": r.plate, "sub": short,
                "ready": bool(r.ready),
                "state": r.state or "pending",
                "reason": r.reason or "",
                "note": r.note or "",
                "arrive_date": r.arrive_date or "",
                "arrive": r.arrive_hhmm or "",
                "location": r.location or "",
                "sheet_status": r.sheet_status or "",
                # A plate nobody committed. Shown on its row rather than only in
                # the import report, which is gone the moment the page reloads.
                "uncommitted": bool(roster) and r.key not in roster,
            })
        # The trucks the counter calls "not accounted": committed for the
        # project and absent from the sheet entirely. They have no row of their
        # own, so without this they are a number with no way to reach the
        # plates - and chasing each one is the job the WI gives the supervisor.
        # Carried as rows so they appear on the board, flagged so that saving
        # the list cannot mistake them for trucks somebody put on it.
        if roster and recs:
            on_sheet = {r.key for r in recs}
            for key, plate in sorted(roster.items()):
                if key in on_sheet:
                    continue
                rows.append({
                    "plate": plate, "sub": short, "ready": False,
                    "state": "absent", "reason": "", "note": "",
                    "arrive_date": "", "arrive": "", "location": "",
                    "sheet_status": "", "uncommitted": False,
                    "absent": True,
                })
    out = {
        "date": day,
        "state": state,
        "rows": rows,
        # Named for what it is: rows sitting with the manager, undecided.
        "applied_count": sum(1 for r in rows if r["state"] == "applied"),
        "submitted_by": dl.submitted_by if dl else None,
        "submitted_at": _fmt(dl.submitted_at) if dl else None,
        "confirmed_by": dl.confirmed_by if dl else None,
        "confirmed_at": _fmt(dl.confirmed_at) if dl else None,
        "reject_reason": (dl.reject_reason or "") if dl else "",
        "role": _role(),
        # The page shows only the buttons whose action is actually available,
        # so every action is asked about here rather than guessed in JavaScript.
        "can": {a: _can(a, state) for a in
                ("edit", "submit", "confirm", "reject", "edit_time", "reopen")},
    }
    if dl:
        s = db.session.get(Subcontractor, dl.subcontractor_id) \
            if dl.subcontractor_id else None
        out["subcontractor_id"] = dl.subcontractor_id
        out["subcontractor"] = s.name if s else None
    return out


# -- Planner settings --------------------------------------------------------
@bp.get("/settings")
@login_required
def get_settings():
    """Everything the planner uses, including the per-leg speeds, in one place."""
    out = []
    for p in PlanSetting.query.order_by(PlanSetting.ordering).all():
        out.append({"key": p.key, "value": p.value, "label": p.label,
                    "unit": p.unit, "group": p.group, "kind": "setting"})
    for r in RouteLeg.query.order_by(RouteLeg.id).all():
        km = 0.0
        if r.points:
            km = sum(engine.haversine_km(r.points[i][0], r.points[i][1],
                                         r.points[i + 1][0], r.points[i + 1][1])
                     for i in range(len(r.points) - 1))
        out.append({"key": "leg:" + r.leg_key, "value": str(r.speed),
                    "label": r.leg_key.replace("_", " → "),
                    "unit": "km/h", "group": "speed", "kind": "leg",
                    "km": round(km, 1),
                    "hours": round(km / r.speed, 2) if r.speed else None})
    return jsonify(settings=out, can_edit=(_role() in ("manager", "admin")))


@bp.post("/settings")
@login_required
def save_settings():
    if _role() not in ("manager", "admin"):
        return jsonify(error="Only a manager or admin may change planning figures"), 403
    d = request.get_json(force=True, silent=True) or {}
    now, who = datetime.utcnow(), current_user.username
    changed = []
    for key, val in (d.get("values") or {}).items():
        val = str(val).strip()
        if key.startswith("leg:"):
            r = RouteLeg.query.filter_by(leg_key=key[4:]).first()
            if not r:
                continue
            try:
                spd = float(val)
            except ValueError:
                return jsonify(error="%s must be a number" % key[4:]), 400
            if spd <= 0:
                return jsonify(error="Speed for %s must be above zero" % key[4:]), 400
            if r.speed != spd:
                changed.append("%s %s -> %s km/h" % (key[4:], r.speed, spd))
                r.speed = spd
        else:
            p = db.session.get(PlanSetting, key)
            if not p:
                continue
            if p.unit == "hours":
                try:
                    if float(val) < 0:
                        raise ValueError
                except ValueError:
                    return jsonify(error="%s must be a number of hours" % p.label), 400
            if p.unit == "time" and not re.match(r"^\d{1,2}:\d{2}$", val):
                return jsonify(error="%s must look like 15:00" % p.label), 400
            if p.value != val:
                changed.append("%s %s -> %s" % (p.label, p.value, val))
                p.value = val
                p.updated_by, p.updated_at = who, now
    db.session.commit()
    return jsonify(ok=True, changed=changed)


@bp.get("/subcontractors")
@login_required
def subcontractors():
    """Companies, with how many trucks each has committed. A subcontractor login
    sees only its own company."""
    mine = getattr(current_user, "subcontractor_id", None)
    q = Subcontractor.query.filter_by(active=True).order_by(Subcontractor.name)
    if _role() == "subcontractor" and mine:
        q = q.filter(Subcontractor.id == mine)
    out = []
    for s in q.all():
        out.append({
            "id": s.id, "name": s.name, "short": s.short or s.name,
            "committed": FleetCommitment.query.filter_by(
                subcontractor_id=s.id, released_on="").count(),
        })
    return jsonify(subcontractors=out, mine=mine, role=_role())


def _roster(sub_id):
    return {c.key: c for c in FleetCommitment.query.filter_by(
        subcontractor_id=sub_id, released_on="").all()}


@bp.post("/upload")
@login_required
def upload():
    """Import a subcontractor's readiness sheet into that day's list.

    Checked against the committed roster BOTH ways: a committed truck absent
    from the sheet is named, and a plate on the sheet that was never committed
    is flagged. Absence is not release - a truck leaves the roster only by an
    explicit, recorded release, never by being left off a sheet.
    """
    import os
    import tempfile

    if _role() not in ("subcontractor", "supervisor", "admin"):
        return jsonify(error="Your role may not upload a readiness sheet"), 403
    f = request.files.get("file")
    if not f:
        return jsonify(error="No file was attached"), 400

    day = (request.form.get("date") or "").strip() \
        or (datetime.utcnow() + LOCAL_OFFSET).strftime("%Y-%m-%d")
    if _asked_all(request.form):
        return _one_company_only()
    sub_id = _req_sub_id(request.form)
    dl = _find_list(day, sub_id)

    # A confirmed list is the document the manager signed. Replacing it from a
    # sheet would rewrite what was approved without anyone deciding to.
    if dl and dl.state == "confirmed":
        s = db.session.get(Subcontractor, sub_id) if sub_id else None
        return jsonify(error="The list for %s on %s is already confirmed. It "
                             "cannot be replaced by an upload."
                             % ((s.name if s else "this company"), day)), 409

    fd, tmp = tempfile.mkstemp(suffix=".xlsx")
    os.close(fd)
    try:
        f.save(tmp)
        parsed = readiness_import.parse(tmp)
    finally:
        try:
            os.remove(tmp)
        except OSError:
            pass

    rows = parsed.get("rows") or []
    if not rows:
        return jsonify(error="No truck rows were found in that sheet.",
                       warnings=parsed.get("warnings", [])), 400

    roster = _roster(sub_id) if sub_id else {}
    on_sheet = {r["key"] for r in rows}
    missing = sorted(c.plate for k, c in roster.items() if k not in on_sheet)
    uncommitted = sorted(r["plate"] for r in rows
                         if roster and r["key"] not in roster)

    if not dl:
        dl = DailyList(list_date=day, subcontractor_id=sub_id, state="draft")
        db.session.add(dl)
        db.session.flush()

    # A truck already sent, or already decided, keeps where it is. Re-uploading a
    # corrected sheet must not drag a decided row back to the start.
    prior = {r.key: (r.state or "pending")
             for r in DailyListRow.query.filter_by(list_id=dl.id).all()}
    DailyListRow.query.filter_by(list_id=dl.id).delete()

    running = 0
    for r in rows:
        runs = readiness_import.is_running(r.get("activity"))
        if runs:
            running += 1
        st = prior.get(r["key"], "pending")
        if st not in ("pending", "applied", "approved", "denied"):
            st = "pending"
        db.session.add(DailyListRow(
            list_id=dl.id, plate=r["plate"], key=r["key"],
            ready=bool(runs), state=st,
            location=(r.get("location") or "")[:60],
            sheet_status=(r.get("status") or "")[:30],
            # The sheet's own words for why it is not running - kept verbatim,
            # because "Maintenace" is what they wrote and what they will ask about.
            reason=("" if runs else (r.get("activity") or "not running")[:300]),
            arrive_date=(r.get("arrive_date") or "")[:10],
            arrive_hhmm=(r.get("arrive_time") or "")[:5],
            note=(r.get("activity") or r.get("remark") or "")[:300]))

    _restate(dl)
    db.session.commit()

    s = db.session.get(Subcontractor, sub_id) if sub_id else None
    return jsonify(
        ok=True, date=day,
        subcontractor=(s.name if s else "(no company)"),
        subcontractor_id=sub_id,
        imported=len(rows), running=running, committed=len(roster),
        missing=missing, uncommitted=uncommitted,
        warnings=parsed.get("warnings", []),
        sheet=parsed.get("sheet"), header_row=parsed.get("header_row"),
        columns=parsed.get("columns"), state=dl.state)


# One stage back, never two. The step back is an undo, not a reset.
BACK_ONE = {"approved": "applied", "denied": "applied", "applied": "pending"}


@bp.post("/rows/revert")
@login_required
def rows_revert():
    """Step named trucks back one stage.

    Whoever made the move is the one who can take it back: a manager undoes an
    approval or a denial, a supervisor withdraws something not yet decided. A
    supervisor cannot quietly undo a manager's decision.
    """
    d = request.get_json(force=True, silent=True) or {}
    day = d.get("date") or (datetime.utcnow() + LOCAL_OFFSET).strftime("%Y-%m-%d")
    if _asked_all(d):
        return _one_company_only()
    sub_id = _req_sub_id(d)
    dl = _find_list(day, sub_id)
    if not dl:
        return jsonify(error="There is no list for this date"), 404

    plates = [str(p).strip() for p in (d.get("plates") or []) if str(p).strip()]
    if not plates:
        return jsonify(error="No trucks were given"), 400
    keys = {engine.norm_plate(p) for p in plates}

    r = _role()
    moved, refused = [], []
    for row in DailyListRow.query.filter_by(list_id=dl.id).all():
        if row.key not in keys:
            continue
        cur = row.state or "pending"
        nxt = BACK_ONE.get(cur)
        if not nxt:
            refused.append("%s is already pending" % row.plate)
            continue
        if cur in ("approved", "denied") and r not in ("manager", "admin"):
            refused.append("%s was decided by the manager" % row.plate)
            continue
        if cur == "applied" and r not in ("supervisor", "manager", "admin"):
            refused.append("%s may only be withdrawn by the supervisor"
                           % row.plate)
            continue
        row.state = nxt
        moved.append({"plate": row.plate, "from": cur, "to": nxt})
    db.session.commit()
    _restate(dl)
    db.session.commit()
    return jsonify(ok=True, moved=moved, refused=refused,
                   state=dl.state, count=len(moved))


def _plan_row(r, label, sub_id, loop, from_plan):
    """One planned loop, in the shape both the week and the revision return.

    The two are compared against each other, so they must be the same object
    with the same field names - a comparison between two slightly different row
    shapes is where a missed factor hides.
    """
    iso = lambda d: d.strftime("%Y-%m-%dT%H:%M") if d else None
    # The QL49 gate wait is the only stop the planner reports as a duration
    # rather than a pair of timestamps; back out when the truck got there.
    ql_wait = r["waits"].get("ql49_gate", 0.0)
    return {
        "plate": r["plate"], "sub": label, "sub_id": sub_id,
        "day": r["arrive_mine"].strftime("%Y-%m-%d"),
        "loop": loop, "from_plan": from_plan,
        "route": r["route"], "cycle_hours": r["cycle_hours"],
        "total_wait": r["total_wait"], "waits": r["waits"],
        "t": {
            "arrive_mine": iso(r["arrive_mine"]),
            "load_start": iso(r["load_start"]), "load_end": iso(r["load_end"]),
            "arrive_border": iso(r["arrive_border"]),
            "cross_border": iso(r["cross_border"]),
            "ql49_arrive": iso(r["ql49_in"] - timedelta(hours=ql_wait)),
            "ql49_in": iso(r["ql49_in"]),
            "arrive_port": iso(r["arrive_port"]),
            "unload_start": iso(r["unload_start"]),
            "unload_end": iso(r["unload_end"]),
            "depart_port": iso(r["depart_port"]),
            "back": iso(r["arrive_mine_back"]),
        },
    }


def _week_start(s):
    """Monday of the ISO week containing s, or of today if s is unusable."""
    try:
        d = datetime.strptime(s, "%Y-%m-%d")
    except (TypeError, ValueError):
        d = datetime.utcnow() + LOCAL_OFFSET
    return (d - timedelta(days=d.weekday())).replace(
        hour=0, minute=0, second=0, microsecond=0)


def _week_data(start_arg, only, roll):
    """Seven days of the plan in one pass.

    The whole window is planned in a SINGLE call, across every company, because
    the bays are shared. Planning one company at a time against two port bays
    describes an operation that does not exist - the queue that decides the
    cycle is formed by whoever is at the gate, not by whoever happens to be on
    one company's list. Planning day by day loses the same way: a truck that
    starts loading at 23:40 holds the bay into the next morning.

    Filtering by company narrows what is SHOWN, never what is queued, so a
    company's times here are the times it will really get.

    Returns a plain dict. The screen and the Excel export both read it, so the
    file a subcontractor is sent cannot drift from the plan on the page.
    """
    start = _week_start(start_arg)
    days = [(start + timedelta(days=i)).strftime("%Y-%m-%d") for i in range(7)]

    # The week is the week trucks REACH THE MINE, not the week the sheets were
    # filed: a list uploaded on Friday routinely carries Saturday's arrivals,
    # and it is the arrival that the plan is built on.
    subs = {s.id: (s.short or s.name) for s in Subcontractor.query.all()}
    owner = {dl.id: (dl.subcontractor_id,
                     subs.get(dl.subcontractor_id, "(no company)"))
             for dl in DailyList.query.all()}

    approved = DailyListRow.query.filter(DailyListRow.state == "approved").all()
    arrivals, who, no_time, touched = [], {}, [], set()
    elsewhere = {}
    for r in approved:
        sub_id, label = owner.get(r.list_id, (None, ""))
        if r.arrive_date and r.arrive_hhmm:
            if r.arrive_date < days[0] or r.arrive_date > days[-1]:
                # Timed, but not this week. Counted so an empty week can say
                # where the trucks actually are instead of just showing nothing.
                if only is None or sub_id == only:
                    elsewhere[r.arrive_date] = elsewhere.get(r.arrive_date, 0) + 1
                continue
            try:
                h, m = r.arrive_hhmm.split(":")
                arrivals.append((r.plate,
                                 datetime.strptime(r.arrive_date, "%Y-%m-%d")
                                 + timedelta(hours=int(h), minutes=int(m))))
                who[r.plate] = (sub_id, label)
                touched.add(r.list_id)
                continue
            except ValueError:
                pass
        # No timing at all: it belongs to whichever week its list was filed for,
        # otherwise it would be invisible in every week.
        dl = db.session.get(DailyList, r.list_id)
        if dl and days[0] <= (dl.list_date or "") <= days[-1] \
                and (only is None or sub_id == only):
            no_time.append(r.plate)
            touched.add(r.list_id)

    # A readiness sheet gets a truck INTO the cycle. After that the plan is what
    # dictates when it goes and when it comes back, and the subcontractor works
    # to the plan rather than the plan being rebuilt from a new sheet every day.
    # So a truck's FIRST arrival is taken from readiness and every later arrival
    # on a sheet is ignored - counted, never silently dropped - while the loops
    # after the first are issued by the plan itself.
    #
    # Only outside events move these times. That correction is the revise pass,
    # a later phase; nothing here pretends to do it.
    first_only, superseded = [], 0
    for plate, dt in sorted(arrivals, key=lambda x: x[1]):
        if any(p == plate for p, _ in first_only):
            superseded += 1
            continue
        first_only.append((plate, dt))
    arrivals = first_only

    cfg = planner.load_config()
    horizon = datetime.strptime(days[-1], "%Y-%m-%d") + timedelta(days=1)
    gap = timedelta(hours=cfg.get("turn_gap_h", 0.0))

    out = planner.plan_trucks(arrivals, cfg) if arrivals else []
    if roll and arrivals:
        extra = []
        for _ in range(8):        # a 45 h cycle fits four loops in a week
            nxt = [(r["plate"], r["arrive_mine_back"] + gap) for r in out
                   if r["arrive_mine_back"] + gap < horizon]
            if len(nxt) == len(extra):
                break
            extra = nxt
            out = planner.plan_trucks(arrivals + extra, cfg)
    committed = set((p, t) for p, t in arrivals)

    rows = []
    per_day = {d: {"cyc": [], "hue": 0, "ql49": 0, "wait": 0.0, "queues": {},
                   "from_plan": 0} for d in days}
    seen = {}
    for r in sorted(out, key=lambda x: x["arrive_mine"]):
        sub_id, label = who.get(r["plate"], (None, ""))
        if only is not None and sub_id != only:
            continue
        seen[r["plate"]] = seen.get(r["plate"], 0) + 1
        from_plan = (r["plate"], r["arrive_mine"]) not in committed
        rows.append(_plan_row(r, label, sub_id, seen[r["plate"]], from_plan))
        b = per_day.get(r["arrive_mine"].strftime("%Y-%m-%d"))
        if b is None:
            continue
        b["cyc"].append(r["cycle_hours"])
        b["hue" if r["route"] == "hue" else "ql49"] += 1
        b["wait"] += r["total_wait"]
        if from_plan:
            b["from_plan"] += 1
        for k, v in r["waits"].items():
            b["queues"][k] = b["queues"].get(k, 0.0) + v

    summary = []
    for d in days:
        b = per_day[d]
        cyc = b["cyc"]
        worst = max(b["queues"].items(), key=lambda kv: kv[1]) if b["queues"] else (None, 0)
        summary.append({
            "date": d, "trucks": len(cyc), "from_plan": b["from_plan"],
            "entering": len(cyc) - b["from_plan"],
            "hue": b["hue"], "ql49": b["ql49"],
            "mean_cycle": round(sum(cyc) / len(cyc), 1) if cyc else None,
            "worst_cycle": round(max(cyc), 1) if cyc else None,
            "over_48": sum(1 for c in cyc if c > 48),
            "wait_hours": round(b["wait"], 1),
            "biggest_queue": worst[0], "biggest_queue_hours": round(worst[1], 1),
        })

    return dict(start=days[0], end=days[-1], days=days, rows=rows,
                summary=summary, no_time=sorted(set(no_time)),
                planned_all=len(out), subcontractor_id=only,
                lists=len(touched), rolled=roll, superseded=superseded,
                turn_gap_hours=cfg.get("turn_gap_h", 0.0),
                elsewhere=[{"date": d, "trucks": n}
                           for d, n in sorted(elsewhere.items())][:6])


@bp.get("/week")
@login_required
def week():
    return jsonify(**_week_data(request.args.get("start"), _req_sub_id(),
                                request.args.get("roll", "1") != "0"))


def _revision_data(day, only):
    """One day, rebuilt from readiness. The revised plan.

    This is the other half of the model. The week is issued in advance and the
    subcontractor works to it; the revision is where readiness takes effect
    again, because by then the operation knows which trucks are actually there
    and when. So it reads the sheet for the day, not the plan.

    It is computed EXACTLY as the week is - same planner, same figures, same
    shared bays across every company - and differs only in what seeds it and in
    stopping after one cycle. Anything else and the difference between the two
    would be an artefact of the arithmetic rather than a fact about the day.
    """
    subs = {s.id: (s.short or s.name) for s in Subcontractor.query.all()}
    owner = {dl.id: (dl.subcontractor_id,
                     subs.get(dl.subcontractor_id, "(no company)"))
             for dl in DailyList.query.all()}

    arrivals, who, no_time = [], {}, []
    elsewhere = {}
    for r in DailyListRow.query.filter(DailyListRow.state == "approved").all():
        sub_id, label = owner.get(r.list_id, (None, ""))
        if r.arrive_date != day:
            # Not this day, but worth knowing about: a revision opened on a
            # quiet day should say where the trucks are, not just draw a blank.
            if r.arrive_date and r.arrive_hhmm and (only is None or sub_id == only):
                elsewhere[r.arrive_date] = elsewhere.get(r.arrive_date, 0) + 1
            continue
        if r.arrive_hhmm:
            try:
                h, m = r.arrive_hhmm.split(":")
                arrivals.append((r.plate, datetime.strptime(day, "%Y-%m-%d")
                                 + timedelta(hours=int(h), minutes=int(m))))
                who[r.plate] = (sub_id, label)
                continue
            except ValueError:
                pass
        if only is None or sub_id == only:
            no_time.append(r.plate)

    cfg = planner.load_config()
    out = planner.plan_trucks(arrivals, cfg) if arrivals else []

    rows, agg = [], {}
    for r in sorted(out, key=lambda x: x["arrive_mine"]):
        sub_id, label = who.get(r["plate"], (None, ""))
        if only is not None and sub_id != only:
            continue
        rows.append(_plan_row(r, label, sub_id, 1, False))
        for k, v in r["waits"].items():
            agg[k] = agg.get(k, 0.0) + v

    cyc = [x["cycle_hours"] for x in rows]
    worst = max(agg.items(), key=lambda kv: kv[1]) if agg else (None, 0)
    return dict(
        date=day, subcontractor_id=only, rows=rows,
        no_time=sorted(set(no_time)), planned_all=len(out),
        dormant=not rows,
        reason=("No approved truck is due at the mine on this day."
                if not arrivals else
                "No truck from this company is due at the mine on this day."),
        elsewhere=[{"date": d, "trucks": n}
                   for d, n in sorted(elsewhere.items())][:6],
        summary={
            "trucks": len(rows),
            "hue": sum(1 for x in rows if x["route"] == "hue"),
            "ql49": sum(1 for x in rows if x["route"] == "ql49"),
            "mean_cycle": round(sum(cyc) / len(cyc), 1) if cyc else None,
            "worst_cycle": round(max(cyc), 1) if cyc else None,
            "over_48": sum(1 for c in cyc if c > 48),
            "wait_hours": round(sum(x["total_wait"] for x in rows), 1),
            "biggest_queue": worst[0],
            "biggest_queue_hours": round(worst[1], 1),
        })


@bp.get("/revision")
@login_required
def revision():
    day = request.args.get("date") or         (datetime.utcnow() + LOCAL_OFFSET).strftime("%Y-%m-%d")
    return jsonify(**_revision_data(day, _req_sub_id()))


def _figures_now():
    """Every planning figure in force, in one flat list, for freezing with a plan."""
    out = [{"key": p.key, "label": p.label, "value": p.value, "unit": p.unit,
            "group": p.group}
           for p in PlanSetting.query.order_by(PlanSetting.ordering).all()]
    for r in RouteLeg.query.order_by(RouteLeg.id).all():
        km = 0.0
        if r.points:
            km = sum(engine.haversine_km(r.points[i][0], r.points[i][1],
                                         r.points[i + 1][0], r.points[i + 1][1])
                     for i in range(len(r.points) - 1))
        out.append({"key": "leg:" + r.leg_key, "label": r.label or r.leg_key,
                    "value": str(r.speed), "unit": "km/h", "group": "speed",
                    "km": round(km, 1)})
    return out


@bp.post("/week/issue")
@login_required
def week_issue():
    """Freeze this week's plan as issued. Only a manager or admin may.

    Issuing is a decision, not a save: from here the subcontractor works to
    these times and the revision is measured against them.
    """
    if _role() not in ("manager", "admin"):
        return jsonify(error="Only a manager or admin may issue a plan"), 403
    d = request.get_json(force=True, silent=True) or {}
    only = _req_sub_id(d)
    data = _week_data(d.get("start"), only, True)
    if not data["rows"]:
        return jsonify(error="There is nothing to issue for this week."), 400

    snap = PlanSnapshot(week_start=data["start"], subcontractor_id=only,
                        issued_by=current_user.username,
                        issued_at=datetime.utcnow(),
                        note=(d.get("note") or "").strip()[:300],
                        rows=data["rows"], figures=_figures_now())
    db.session.add(snap)
    db.session.commit()
    return jsonify(ok=True, id=snap.id, week_start=snap.week_start,
                   issued_by=snap.issued_by, issued_at=_fmt(snap.issued_at),
                   loops=len(data["rows"]))


def _issued_for(day, only):
    """The plan in force for a day: the newest snapshot whose week contains it.

    A company's own issued plan wins over the all-companies one; without either
    there is no promise to compare against, and this says so rather than
    inventing a baseline.
    """
    week = _week_start(day).strftime("%Y-%m-%d")
    q = PlanSnapshot.query.filter_by(week_start=week)
    mine = q.filter_by(subcontractor_id=only).order_by(
        PlanSnapshot.issued_at.desc()).first() if only is not None else None
    return mine or q.filter_by(subcontractor_id=None).order_by(
        PlanSnapshot.issued_at.desc()).first()


@bp.get("/week/issued")
@login_required
def week_issued():
    """What has been issued for a week, newest first."""
    start = _week_start(request.args.get("start")).strftime("%Y-%m-%d")
    out = []
    for s in PlanSnapshot.query.filter_by(week_start=start).order_by(
            PlanSnapshot.issued_at.desc()).all():
        sub = db.session.get(Subcontractor, s.subcontractor_id)             if s.subcontractor_id else None
        out.append({"id": s.id, "week_start": s.week_start,
                    "subcontractor_id": s.subcontractor_id,
                    "subcontractor": (sub.short or sub.name) if sub else "All companies",
                    "issued_by": s.issued_by, "issued_at": _fmt(s.issued_at),
                    "loops": len(s.rows or []), "note": s.note or ""})
    return jsonify(week_start=start, issued=out)


# The stops worth comparing. A shift at the mine that the corridor absorbs is
# noise; a shift still there at the port is the day costing you a cycle.
VARIANCE_POINTS = [
    ("arrive_mine", "at the mine"),
    ("cross_border", "crossing the border"),
    ("arrive_port", "at the port"),
    ("back", "back at the mine"),
]


def _mins(a, b):
    """b - a in whole minutes, or None if either side is missing."""
    if not a or not b:
        return None
    return int(round((datetime.strptime(b, "%Y-%m-%dT%H:%M")
                      - datetime.strptime(a, "%Y-%m-%dT%H:%M")).total_seconds() / 60))


@bp.get("/variance")
@login_required
def variance():
    """The revised day against the week that was issued.

    The week is the promise; the revision is the day as readiness now describes
    it. The difference between them is the list of factors the plan did not know
    about - which is the only reason to keep the two apart.

    The baseline is the ISSUED plan, read back from the snapshot. It cannot be
    recomputed here: the week is built from the same readiness rows the revision
    reads, so recomputing it would move the promise by exactly the amount the day
    moved and every difference would cancel to zero - a comparison that reports
    "as planned" however badly the day went. Measured against a frozen plan, a
    truck that moved, moved.
    """
    day = request.args.get("date") or         (datetime.utcnow() + LOCAL_OFFSET).strftime("%Y-%m-%d")
    only = _req_sub_id()

    rev = _revision_data(day, only)
    snap = _issued_for(day, only)
    if snap is None:
        return jsonify(
            date=day, subcontractor_id=only, rows=[], baseline=None,
            reason="No plan has been issued for this week, so there is nothing to "
                   "measure the day against. Issue the week first, on the Week tab.",
            summary={})

    sub = db.session.get(Subcontractor, snap.subcontractor_id)         if snap.subcontractor_id else None
    baseline = {"id": snap.id, "week_start": snap.week_start,
                "issued_by": snap.issued_by, "issued_at": _fmt(snap.issued_at),
                "scope": (sub.short or sub.name) if sub else "All companies",
                "loops": len(snap.rows or [])}

    # The baseline loop for a truck is the one it was to start THAT day.
    planned = {}
    for r in (snap.rows or []):
        if r.get("day") == day and (only is None or r.get("sub_id") == only):
            planned.setdefault(r["plate"], r)

    rows, agg = [], {}
    for r in rev["rows"]:
        p = planned.pop(r["plate"], None)
        if not p:
            rows.append({"plate": r["plate"], "sub": r["sub"], "status": "unplanned",
                         "revised": r["t"], "planned": None, "delta": {},
                         "cycle_delta": None, "cycle_hours": r["cycle_hours"],
                         "waits": {}})
            continue
        delta = {k: _mins(p["t"].get(k), r["t"].get(k)) for k, _ in VARIANCE_POINTS}
        waits = {}
        for k in set(list(p["waits"].keys()) + list(r["waits"].keys())):
            d = round(r["waits"].get(k, 0.0) - p["waits"].get(k, 0.0), 2)
            if d:
                waits[k] = d
                agg[k] = agg.get(k, 0.0) + d
        rows.append({
            "plate": r["plate"], "sub": r["sub"],
            "status": "moved" if any(delta.values()) else "as planned",
            "planned": p["t"], "revised": r["t"], "delta": delta,
            "cycle_hours": r["cycle_hours"],
            "cycle_delta": round(r["cycle_hours"] - p["cycle_hours"], 1),
            "waits": waits,
        })

    # Whatever is left was in the plan and is not on the day's sheet.
    for plate, p in planned.items():
        rows.append({"plate": plate, "sub": p["sub"], "status": "missing",
                     "planned": p["t"], "revised": None, "delta": {},
                     "cycle_delta": None, "cycle_hours": None, "waits": {}})

    back = [r["delta"].get("back") for r in rows
            if r["status"] == "moved" and r["delta"].get("back") is not None]
    worst = max(agg.items(), key=lambda kv: abs(kv[1])) if agg else (None, 0)
    moved = [r for r in rows if r["status"] == "moved"]
    return jsonify(
        date=day, subcontractor_id=only, rows=rows, baseline=baseline,
        summary={
            "compared": sum(1 for r in rows if r["status"] in ("moved", "as planned")),
            "moved": len(moved),
            "as_planned": sum(1 for r in rows if r["status"] == "as planned"),
            "unplanned": sum(1 for r in rows if r["status"] == "unplanned"),
            "missing": sum(1 for r in rows if r["status"] == "missing"),
            "mean_back": int(round(sum(back) / len(back))) if back else 0,
            "worst_back": (max(back, key=abs) if back else 0),
            # Named rather than left to be inferred: this is the answer to
            # "what did we miss", and it should not need arithmetic to read.
            "factor": (worst[0] or "").replace("_", " "),
            "factor_hours": round(worst[1], 1),
        })


# Until the real driver numbers are loaded against each truck (Truck.phone, set
# on the fleet page), every row falls back to this one. Marked as a placeholder
# in the payload so the screen can say so rather than implying it reached a driver.
DEFAULT_DRIVER_PHONE = "+66817916147"

# A corridor timed in gates and queues does not run to the minute. Inside an hour
# either side of the plan a truck is running to plan; outside it, it is early or
# it is late, and somebody should know which.
ON_TIME_MINUTES = 60


# The corridor as four places, seen twice: loaded on the way out, empty on the
# way home. Each entry is (key, label, the plan field that promises it, the
# geofence role that proves it, which edge of the visit counts).
LOC_FH = [
    ("mine",   "Mine",   "arrive_mine",   "xppl",   "enter"),
    ("border", "Border", "arrive_border", "border", "enter"),
    ("ql49",   "QL49",   "ql49_in",       "ql49",   "enter"),
    ("port",   "Port",   "arrive_port",   "port",   "enter"),
]
# Going home the plan times only the two ends. The middle two are real places a
# truck passes and GPS sees, so they are shown with an actual and no promise
# rather than left off the board.
LOC_BH = [
    ("port",   "Port",   "depart_port", "port",   "exit"),
    ("ql49",   "QL49",   None,          "ql49",   "enter"),
    ("border", "Border", None,          "border", "enter"),
    ("mine",   "Mine",   "back",        "xppl",   "enter"),
]


@bp.get("/track")
@login_required
def track():
    """Every truck, place by place: what the plan said and what happened.

    Where a place has not been reached yet the actual is an ESTIMATE, marked as
    one: the plan carried forward by however late the truck already is, or - for
    a truck still on the road with nothing confirmed - the distance left to the
    mine over the planner's own empty-running speed.
    """
    day = request.args.get("date") or \
        (datetime.utcnow() + LOCAL_OFFSET).strftime("%Y-%m-%d")
    only = _req_sub_id()

    snap = _issued_for(day, only)
    if snap:
        plan_rows = [r for r in (snap.rows or [])
                     if only is None or r.get("sub_id") == only]
        source = "issued %s" % _fmt(snap.issued_at)
    else:
        plan_rows = _revision_data(day, only)["rows"]
        source = "today's revision - no plan has been issued for this week"

    # A sheet is filed the day BEFORE the trucks run, so the board's date and the
    # loop's date are rarely the same. Match a truck to the first loop it starts
    # on or after this date.
    by_plate = {}
    for r in sorted(plan_rows, key=lambda x: x["t"]["arrive_mine"]):
        if r["day"] < day:
            continue
        by_plate.setdefault(r["plate"], r)

    dl = _find_list(day, only)
    listed = []
    if dl:
        listed = [r.plate for r in DailyListRow.query.filter_by(
            list_id=dl.id, state="approved").order_by(DailyListRow.plate).all()]
    sub = db.session.get(Subcontractor, only) if only else None
    sub_short = (sub.short or sub.name) if sub else ""

    phones = {engine.norm_plate(t.plate): (t.phone or "").strip()
              for t in Truck.query.all()}

    visits, roles = _visits_and_roles()
    lo, _hi = _day_bounds(day)
    seen_anchor = {v["anchor_id"] for v in visits}
    by_plate_visits = {}
    for v in visits:
        if v["enter"] < lo or v["enter"] > lo + CYCLE_SPAN:
            continue
        by_plate_visits.setdefault(engine.norm_plate(v["plate"]), []).append(v)

    last_ping = {}
    for g in GpsPing.query.filter(GpsPing.dt >= lo,
                                  GpsPing.dt <= lo + CYCLE_SPAN).order_by(GpsPing.dt).all():
        last_ping[engine.norm_plate(g.plate)] = g

    home = RouteLeg.query.filter_by(leg_key="mine_border").first()
    to_mine = list(reversed(home.points or [])) if home else []
    empty = RouteLeg.query.filter_by(leg_key="port_mine").first()
    empty_kmh = (empty.speed if empty and empty.speed else 40.0)

    iso = lambda d: _local(d).strftime("%Y-%m-%dT%H:%M") if d else None
    blind = set()
    for _k, _l, _p, role, _e in LOC_FH + LOC_BH:
        aid = roles.get(role)
        if aid is None or aid not in seen_anchor:
            blind.add(role)

    rows = []
    for plate in sorted(set(list(by_plate.keys()) + listed)):
        p = by_plate.get(plate)
        plan = (p or {}).get("t") or {}
        vs = sorted(by_plate_visits.get(engine.norm_plate(plate), []),
                    key=lambda x: x["enter"])

        # The port splits the loop: everything before it is the loaded run, and
        # a second pass at the same place afterwards is the run home.
        port_id = roles.get("port")
        split = None
        for v in vs:
            if v["anchor_id"] == port_id:
                split = v
                break

        def pick(role, edge, leg):
            aid = roles.get(role)
            if not aid:
                return None
            for v in vs:
                if v["anchor_id"] != aid:
                    continue
                if leg == "fh" and split is not None and v["enter"] > split["enter"]:
                    continue
                if leg == "bh":
                    if split is None:
                        continue
                    if role != "port" and v["enter"] <= (split.get("exit") or split["enter"]):
                        continue
                return v.get(edge) or v.get("enter")
            return None

        last_seen = None
        for key, label, field, role, edge in LOC_FH + LOC_BH:
            leg = "fh" if (key, label, field, role, edge) in LOC_FH else "bh"
            got = pick(role, edge, leg)
            if got:
                last_seen = {"place": label, "leg": leg, "at": iso(got)}

        # Nothing confirmed yet, but pinging on the road: the distance left says
        # more than the plan does.
        enroute = None
        if last_seen is None:
            g = last_ping.get(engine.norm_plate(plate))
            if g is not None and to_mine:
                km = engine.remaining_km_along_route(to_mine, g.lat, g.lng)
                if km is not None:
                    enroute = {"at": iso(g.dt), "km_to_mine": round(km, 1),
                               "eta": iso(g.dt + timedelta(hours=km / empty_kmh))}

        # Walked in the order the truck runs, carrying the delay forward: an hour
        # lost at the border is still lost at the port. Confirmed times reset it,
        # because the corridor's gates can absorb a delay as easily as create one.
        drift = [0]

        def cells(spec, leg):
            out = []
            for key, label, field, role, edge in spec:
                got = pick(role, edge, leg)
                planned = plan.get(field) if field else None
                cell = {"key": key, "label": label, "plan": planned,
                        "actual": iso(got), "estimate": None, "delay": None,
                        "blind": role in blind}
                if got and planned:
                    d = _mins(planned, iso(got))
                    cell["delay"] = d
                    if d is not None:
                        drift[0] = d
                elif not got and planned:
                    if enroute and leg == "fh" and key == "mine":
                        cell["estimate"] = enroute["eta"]
                        cell["delay"] = _mins(planned, enroute["eta"])
                        if cell["delay"] is not None:
                            drift[0] = cell["delay"]
                    else:
                        cell["estimate"] = (datetime.strptime(planned, "%Y-%m-%dT%H:%M")
                                            + timedelta(minutes=drift[0])
                                            ).strftime("%Y-%m-%dT%H:%M")
                        cell["delay"] = drift[0]
                out.append(cell)
            return out

        # Computed in journey order, returned in COLUMN order, so the same four
        # columns hold the same four places whichever leg is on screen.
        def by_column(spec_cells):
            ix = {c["key"]: c for c in spec_cells}
            return [ix[k] for k, _l, _f, _r, _e in LOC_FH]

        phone = phones.get(engine.norm_plate(plate)) or ""
        rows.append({
            "plate": plate, "sub": (p or {}).get("sub") or sub_short,
            "phone": phone or DEFAULT_DRIVER_PHONE,
            "phone_known": bool(phone),
            "planned": bool(p), "route": (p or {}).get("route"),
            "cycle_hours": (p or {}).get("cycle_hours"),
            "fh": by_column(cells(LOC_FH, "fh")),
            "bh": by_column(cells(LOC_BH, "bh")),
            "last_seen": last_seen, "enroute": enroute, "drift": drift[0],
        })

    moving = [r for r in rows if r["last_seen"] or r["enroute"]]
    running_late = [r for r in rows if (r["drift"] or 0) > ON_TIME_MINUTES]
    running_early = [r for r in rows if (r["drift"] or 0) < -ON_TIME_MINUTES]
    return jsonify(
        date=day, subcontractor_id=only, source=source,
        locations=[{"key": k, "label": l} for k, l, _f, _r, _e in LOC_FH],
        blind=sorted(blind), rows=rows, on_time_minutes=ON_TIME_MINUTES,
        summary={
            "trucks": len(rows),
            "planned": sum(1 for r in rows if r["planned"]),
            "moving": len(moving),
            "on_the_road": sum(1 for r in rows if r["enroute"]),
            "not_seen": sum(1 for r in rows if not r["last_seen"] and not r["enroute"]),
            "late": len(running_late),
            "early": len(running_early),
            "worst_late": max([r["drift"] for r in running_late] or [0]),
            "home": sum(1 for r in rows if r["bh"][0]["actual"]),
        })


WEEK_COLS = [
    ("Company", "sub", 14), ("Truck", "plate", 13), ("Loop", "loop", 6),
    ("Dispatched by", "source", 14),
    ("Arrive mine", "arrive_mine", 15), ("Load start", "load_start", 15),
    ("Load end", "load_end", 15), ("Arrive border", "arrive_border", 15),
    ("Cross border", "cross_border", 15), ("Reach QL49", "ql49_arrive", 15),
    ("Enter QL49", "ql49_in", 15), ("Arrive port", "arrive_port", 15),
    ("Unload start", "unload_start", 15), ("Unload end", "unload_end", 15),
    ("Depart port", "depart_port", 15), ("Back at mine", "back", 15),
    ("Road home", "route", 11), ("Cycle (h)", "cycle_hours", 10),
    ("Waiting (h)", "total_wait", 11),
    ("Wait: mine bay", "mine_queue", 14), ("Wait: border", "border_gate", 13),
    ("Wait: QL49 in", "ql49_gate", 13), ("Wait: port", "port_queue", 12),
    ("Wait: QL49 out", "ql49_out_gate", 14),
]


@bp.get("/week.xlsx")
@login_required
def week_export():
    """The week as a workbook: the plan, the day totals, and the figures it was
    computed from.

    The figures travel WITH the plan on purpose. This file is what a company is
    sent and what a revised plan is later compared against, and a comparison
    against a plan whose assumptions were not written down cannot say which
    factor moved.
    """
    import io
    import openpyxl
    from openpyxl.styles import Font, Alignment

    only = _req_sub_id()
    d = _week_data(request.args.get("start"), only,
                   request.args.get("roll", "1") != "0")
    who = "All companies"
    if only:
        s = db.session.get(Subcontractor, only)
        who = (s.short or s.name) if s else str(only)

    dt = lambda v: datetime.strptime(v, "%Y-%m-%dT%H:%M") if v else None
    head = Font(bold=True)
    wb = openpyxl.Workbook()

    ws = wb.active
    ws.title = "Plan"
    ws.append(["Dispatch plan  %s to %s  -  %s  -  local time (UTC+7)"
               % (d["start"], d["end"], who)])
    ws["A1"].font = Font(bold=True, size=13)
    ws.append([])
    ws.append([c[0] for c in WEEK_COLS])
    for c in ws[3]:
        c.font = head
        c.alignment = Alignment(horizontal="center")
    for r in d["rows"]:
        line = []
        for _, key, _w in WEEK_COLS:
            if key == "source":
                line.append("plan" if r["from_plan"] else "readiness")
            elif key == "route":
                line.append("Hue" if r["route"] == "hue" else "QL49")
            elif key in r:
                line.append(r[key])
            elif key in r["t"]:
                line.append(dt(r["t"][key]))
            else:
                # An explicit 0 rather than a blank: this file gets diffed
                # against a revised plan, and "no wait" must not read as
                # "not recorded".
                line.append(round(r["waits"].get(key, 0.0), 2))
        ws.append(line)
    for i, (_lbl, key, w) in enumerate(WEEK_COLS, 1):
        col = openpyxl.utils.get_column_letter(i)
        ws.column_dimensions[col].width = w
        if key in ("arrive_mine", "load_start", "load_end", "arrive_border",
                   "cross_border", "ql49_arrive", "ql49_in", "arrive_port",
                   "unload_start", "unload_end", "depart_port", "back"):
            for cell in ws[col][3:]:
                cell.number_format = "ddd dd/mm hh:mm"
    ws.freeze_panes = "E4"
    ws.auto_filter.ref = "A3:%s%d" % (
        openpyxl.utils.get_column_letter(len(WEEK_COLS)), 3 + len(d["rows"]))

    ws2 = wb.create_sheet("By day")
    ws2.append(["Date", "Day", "Trucks", "First dispatch", "Issued by plan",
                "Mean cycle (h)", "Worst (h)", "Over 48 h", "Via Hue", "Via QL49",
                "Waiting (h)", "Most of it"])
    for c in ws2[1]:
        c.font = head
    for s in d["summary"]:
        day = datetime.strptime(s["date"], "%Y-%m-%d")
        ws2.append([s["date"], day.strftime("%a"), s["trucks"], s["entering"],
                    s["from_plan"], s["mean_cycle"], s["worst_cycle"],
                    s["over_48"], s["hue"], s["ql49"], s["wait_hours"],
                    (s["biggest_queue"] or "").replace("_", " ")])
    for col, w in zip("ABCDEFGHIJKL", (12, 6, 8, 14, 14, 14, 10, 10, 9, 9, 12, 16)):
        ws2.column_dimensions[col].width = w
    ws2.freeze_panes = "A2"

    ws3 = wb.create_sheet("Figures used")
    ws3.append(["Every number below was in force when this plan was made. A "
                "revised plan is compared against these, not against today's."])
    ws3["A1"].font = Font(italic=True)
    ws3.append([])
    ws3.append(["Figure", "Value", "Unit", "Group"])
    for c in ws3[3]:
        c.font = head
    for p in PlanSetting.query.order_by(PlanSetting.ordering).all():
        ws3.append([p.label, p.value, p.unit, p.group])
    ws3.append([])
    ws3.append(["Leg", "km", "km/h", "hours"])
    for c in ws3[ws3.max_row]:
        c.font = head
    for r in RouteLeg.query.order_by(RouteLeg.id).all():
        km = 0.0
        if r.points:
            km = sum(engine.haversine_km(r.points[i][0], r.points[i][1],
                                         r.points[i + 1][0], r.points[i + 1][1])
                     for i in range(len(r.points) - 1))
        ws3.append([r.label or r.leg_key, round(km, 1), r.speed,
                    round(km / r.speed, 2) if r.speed else None])
    ws3.append([])
    ws3.append(["Loops issued by the plan included",
                "yes" if d["rolled"] else "no (first dispatch only)"])
    ws3.append(["Later readiness times ignored", d["superseded"]])
    ws3.append(["Generated", (datetime.utcnow() + LOCAL_OFFSET).strftime("%Y-%m-%d %H:%M")])
    ws3.append(["Generated by", current_user.username])
    for col, w in zip("ABCD", (42, 14, 10, 12)):
        ws3.column_dimensions[col].width = w

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    name = "dispatch_plan_%s_%s.xlsx" % (
        d["start"], who.replace(" ", "-").replace("/", "-"))
    return Response(buf.read(),
                    mimetype="application/vnd.openxmlformats-officedocument."
                             "spreadsheetml.sheet",
                    headers={"Content-Disposition": 'attachment; filename="%s"' % name})


@bp.get("/summary")
@login_required
def summary():
    """Every company's list for one day, in one place - what a manager needs to
    see before approving anything."""
    day = request.args.get("date") or (datetime.utcnow() + LOCAL_OFFSET).strftime("%Y-%m-%d")
    only = _req_sub_id()      # follow the picker: the total covers what is shown
    out, tot = [], {"committed": 0, "trucks": 0, "sent": 0, "pending": 0}
    q = Subcontractor.query.filter_by(active=True)
    if only:
        q = q.filter(Subcontractor.id == only)
    for sub in q.order_by(Subcontractor.name).all():
        committed = FleetCommitment.query.filter_by(
            subcontractor_id=sub.id, released_on="").count()
        dl = _find_list(day, sub.id)
        rows = DailyListRow.query.filter_by(list_id=dl.id).all() if dl else []
        sent = sum(1 for r in rows if (r.state or "") == "approved")
        pend = sum(1 for r in rows if (r.state or "") == "pending")
        if not committed and not rows and only != sub.id:
            continue                     # company not in play - do not list it
        missing = 0
        if committed and rows:
            keys = {r.key for r in rows}
            missing = FleetCommitment.query.filter(
                FleetCommitment.subcontractor_id == sub.id,
                FleetCommitment.released_on == "",
                ~FleetCommitment.key.in_(keys)).count()
        out.append({"id": sub.id, "name": sub.name, "short": sub.short or sub.name,
                    "committed": committed, "state": (dl.state if dl else "none"),
                    "trucks": len(rows), "sent": sent, "pending": pend,
                    "missing": missing,
                    "confirmed_at": _fmt(dl.confirmed_at) if dl else None})
        tot["committed"] += committed; tot["trucks"] += len(rows)
        tot["sent"] += sent; tot["pending"] += pend
    return jsonify(date=day, parties=out, total=tot, role=_role())

@bp.get("/latest")
@login_required
def latest():
    """The date the page should open on.

    Today is usually empty - a list is built the day before it runs - so opening
    on today shows nothing. Prefer the most recently submitted list, since that
    is the day someone is actually working on; fall back to the newest list of
    any state, then to today.
    """
    today = (datetime.utcnow() + LOCAL_OFFSET).strftime("%Y-%m-%d")
    q = DailyList.query
    if _role() == "subcontractor":
        q = q.filter_by(subcontractor_id=getattr(current_user, "subcontractor_id", None))
    # Sheets arrive daily, so today's sheet is normally the one being worked.
    # Fall back to the most recent only when today has nothing yet.
    t = q.filter_by(list_date=today).order_by(DailyList.id.desc()).first()
    if t:
        return jsonify(date=t.list_date, subcontractor_id=t.subcontractor_id,
                       why="today", today=today)
    sub = (q.filter(DailyList.submitted_at.isnot(None))
           .order_by(DailyList.submitted_at.desc()).first())
    if sub:
        return jsonify(date=sub.list_date, subcontractor_id=sub.subcontractor_id,
                       why="most recently submitted", today=today)
    any_list = q.order_by(DailyList.list_date.desc()).first()
    if any_list:
        return jsonify(date=any_list.list_date, subcontractor_id=any_list.subcontractor_id,
                       why="newest list", today=today)
    return jsonify(date=today, subcontractor_id=None, why="no lists yet", today=today)


@bp.get("/alerts")
@login_required
def alerts():
    """What has happened that this person has not seen yet.

    Only confirming raises one. A submit moves the list from the supervisor to
    the manager and leaves every row `applied`, which is not yet plannable;
    confirming is what turns rows `approved`, and approved is what the planner
    can actually build a day from. Alerting on submit would be telling the
    planner about work that is not theirs to do yet.

    Never raised for whoever did it - they were there - and a subcontractor
    login only ever hears about its own company.
    """
    seen = getattr(current_user, "alerts_seen_at", None)
    # A week back at most. A login left unused for a month should open on a
    # clean board, not on a month of history it can do nothing about.
    floor = datetime.utcnow() - timedelta(days=7)
    since = max(seen, floor) if seen else floor

    subs = {x.id: (x.short or x.name) for x in Subcontractor.query.all()}
    mine = getattr(current_user, "subcontractor_id", None) if _role() == "subcontractor" else None

    out = []
    q = DailyList.query.filter(DailyList.confirmed_at.isnot(None))
    for dl in q.order_by(DailyList.confirmed_at.desc()).limit(50).all():
        if dl.confirmed_at is None or dl.confirmed_at <= since:
            continue
        if mine is not None and dl.subcontractor_id != mine:
            continue
        if (dl.confirmed_by or "") == current_user.username:
            continue
        rows = DailyListRow.query.filter_by(list_id=dl.id, state="approved").all()
        due = {}
        for r in rows:
            if r.arrive_date and r.arrive_hhmm:
                due[r.arrive_date] = due.get(r.arrive_date, 0) + 1
        out.append({
            "kind": "confirmed",
            "sheet_date": dl.list_date,
            "company": subs.get(dl.subcontractor_id, "(no company)"),
            "subcontractor_id": dl.subcontractor_id,
            "by": dl.confirmed_by,
            "at": _fmt(dl.confirmed_at),
            "approved": len(rows),
            # The day the planner cares about, which is not the sheet's date.
            "due": [{"date": d, "trucks": n} for d, n in sorted(due.items())],
            "no_time": sum(1 for r in rows if not (r.arrive_date and r.arrive_hhmm)),
        })
    return jsonify(alerts=out, count=len(out))


@bp.post("/alerts/seen")
@login_required
def alerts_seen():
    """Clear this person's alerts. Their own mark only."""
    current_user.alerts_seen_at = datetime.utcnow()
    db.session.commit()
    return jsonify(ok=True, at=_fmt(current_user.alerts_seen_at))


@bp.get("/plan-day")
@login_required
def plan_day():
    """The day at the mine the planner should show.

    The planner asks a different question from the readiness board, and the two
    answers are normally a day apart. The board works on the sheet being built,
    which is today's; the planner works on the day those trucks actually reach
    the mine, which is tomorrow. Opening the planner on the board's date drew an
    empty plan while a full sheet sat one day ahead - which reads, fairly, as
    "the manager submitted the list and the planner cannot see it".

    So the day comes from the arrivals themselves: today if trucks are due
    today, else the next day that has any, else the most recent day that did -
    a quiet morning still opens on the run just finished rather than on nothing.
    """
    today = (datetime.utcnow() + LOCAL_OFFSET).strftime("%Y-%m-%d")
    only = _req_sub_id()

    # Only approved rows carry a promise; a pending truck is not due anywhere.
    owner = {dl.id: dl.subcontractor_id for dl in DailyList.query.all()}
    days = {}
    for r in (DailyListRow.query
              .filter(DailyListRow.state == "approved").all()):
        if not r.arrive_date or not r.arrive_hhmm:
            continue
        if only is not None and owner.get(r.list_id) != only:
            continue
        days[r.arrive_date] = days.get(r.arrive_date, 0) + 1

    if not days:
        return jsonify(date=today, trucks=0, why="nothing due", today=today,
                       days=[])
    if today in days:
        day, why = today, "due today"
    else:
        ahead = sorted(d for d in days if d > today)
        if ahead:
            day, why = ahead[0], "next day with trucks due"
        else:
            day, why = max(days), "most recent day with trucks"
    return jsonify(date=day, trucks=days[day], why=why, today=today,
                   days=[{"date": d, "trucks": n}
                         for d, n in sorted(days.items())])


@bp.get("/list")
@login_required
def get_list():
    day = request.args.get("date") or datetime.utcnow().strftime("%Y-%m-%d")
    if _asked_all():
        return jsonify(_all_payload(day))
    sub_id = _req_sub_id()
    dl = _find_list(day, sub_id)
    out = _list_payload(dl, day)
    if dl is None:
        out["subcontractor_id"] = sub_id
        s = db.session.get(Subcontractor, sub_id) if sub_id else None
        out["subcontractor"] = s.name if s else None
    return jsonify(out)


@bp.post("/list")
@login_required
def save_list():
    """Replace the day's rows. Allowed while draft/rejected; once confirmed only
    a manager or admin may change it."""
    d = request.get_json(force=True, silent=True) or {}
    day = d.get("date") or datetime.utcnow().strftime("%Y-%m-%d")
    if _asked_all(d):
        return _one_company_only()
    sub_id = _req_sub_id(d)
    dl = _find_list(day, sub_id)
    state = dl.state if dl else "none"
    if not _can("edit", state):
        return jsonify(error="Your role may not edit this list (state: %s)" % state), 403

    if not dl:
        dl = DailyList(list_date=day, subcontractor_id=sub_id, state="draft")
        db.session.add(dl)
        db.session.flush()

    incoming = [r for r in (d.get("rows") or []) if (r.get("plate") or "").strip()]

    # Keep the state a row already has: a row sitting with the manager, or already
    # decided, must not be dragged back to pending because the supervisor saved
    # the page. The tick is intent; the state is where the row actually is.
    prior = {r.key: (r.state or "pending")
             for r in DailyListRow.query.filter_by(list_id=dl.id).all()}
    DailyListRow.query.filter_by(list_id=dl.id).delete()
    for r in incoming:
        plate = (r.get("plate") or "").strip()
        ready = bool(r.get("ready", True))
        st = prior.get(engine.norm_plate(plate), "pending")
        if st not in ("approved", "pending", "denied", "applied"):
            st = "pending"
        db.session.add(DailyListRow(list_id=dl.id, plate=plate,
                                    key=engine.norm_plate(plate),
                                    ready=ready, state=st,
                                    location=(r.get("location") or "").strip()[:60],
                                    sheet_status=(r.get("sheet_status") or "").strip()[:30],
                                    reason=(r.get("reason") or "").strip(),
                                    arrive_date=(r.get("arrive_date") or "").strip()[:10],
                                    arrive_hhmm=(r.get("arrive") or "").strip()[:5],
                                    note=(r.get("note") or "").strip()))
    db.session.commit()
    return jsonify(_list_payload(dl, day))


@bp.post("/list/<action>")
@login_required
def act_list(action):
    if action not in ("submit", "confirm", "reject", "reopen"):
        return jsonify(error="unknown action"), 404
    d = request.get_json(force=True, silent=True) or {}
    day = d.get("date") or datetime.utcnow().strftime("%Y-%m-%d")
    if _asked_all(d):
        return _one_company_only()
    sub_id = _req_sub_id(d)
    dl = _find_list(day, sub_id)
    state = dl.state if dl else "none"
    if not _can(action, state):
        return jsonify(error="Your role (%s) may not %s a list in state '%s'"
                             % (_role(), action, state)), 403
    if not dl:
        return jsonify(error="no list for %s" % day), 404
    if action == "submit":
        rows = DailyListRow.query.filter_by(list_id=dl.id).all()
        # Only the ticked rows go. Trucks still being chased stay with the
        # supervisor and can be sent later - one unresolved truck must not hold
        # up the thirty that are ready.
        moving = [r for r in rows if r.ready and (r.state or "pending") == "pending"]
        if not moving:
            return jsonify(error="No new trucks are ticked. Tick the trucks you are "
                                 "sending to the manager."), 400

    who, now = current_user.username, datetime.utcnow()
    moved = 0
    if action == "submit":
        for r in moving:
            r.state = "applied"
            moved += 1
        dl.submitted_by, dl.submitted_at = who, now
        dl.reject_reason = ""
    elif action == "confirm":
        # Decide every row currently sitting with the manager: ticked is
        # approved, unticked is denied. Rows the supervisor has not sent are
        # untouched - they are not the manager's to decide yet.
        for r in DailyListRow.query.filter_by(list_id=dl.id, state="applied").all():
            r.state = "approved" if r.ready else "denied"
            moved += 1
        if not moved:
            return jsonify(error="Nothing is waiting for approval."), 400
        dl.confirmed_by, dl.confirmed_at = who, now
    elif action == "reject":
        reason = (d.get("reason") or "").strip()
        if not reason:
            return jsonify(error="a reason is required to reject"), 400
        for r in DailyListRow.query.filter_by(list_id=dl.id, state="applied").all():
            r.state = "pending"          # straight back to the supervisor
            moved += 1
        dl.rejected_by, dl.rejected_at = who, now
        dl.reject_reason = reason
    db.session.commit()
    _restate(dl)
    db.session.commit()
    return jsonify(_list_payload(dl, day))


@bp.get("/config")
@login_required
def config():
    out = []
    for s in Shift.query.filter_by(active=True).order_by(Shift.ordering).all():
        checks = ShiftCheck.query.filter_by(shift_id=s.id).order_by(ShiftCheck.ordering).all()
        out.append({
            "id": s.id, "name": s.name,
            "start": s.start_hhmm, "end": s.end_hhmm,
            "checks": [{"code": c.code, "label": c.label} for c in checks],
        })
    return jsonify(shifts=out, role=getattr(current_user, "role", "monitor"))


@bp.get("/board")
@login_required
def board():
    """For one shift on one day: per check, which trucks are done and which are not."""
    day = request.args.get("date") or datetime.utcnow().strftime("%Y-%m-%d")
    try:
        lo, hi = _day_bounds(day)
    except ValueError:
        return jsonify(error="bad date"), 400

    shift = None
    sid = request.args.get("shift_id", type=int)
    if sid:
        shift = db.session.get(Shift, sid)
    if shift is None:
        shift = Shift.query.filter_by(active=True).order_by(Shift.ordering).first()
    if shift is None:
        return jsonify(error="no shifts configured"), 404

    checks = ShiftCheck.query.filter_by(shift_id=shift.id).order_by(ShiftCheck.ordering).all()
    sub_id = _req_sub_id()
    dl = _find_list(day, sub_id)

    # The confirmed list is what opens the cycles. Without one there is nothing to
    # expect, so the board is dormant - it does not invent work.
    if dl is None or dl.state != "confirmed":
        why = {
            "none": "No truck list has been made for this date.",
            "draft": "The list for this date is still a draft - the supervisor has "
                     "not submitted it yet.",
            "submitted": "The list has been submitted but the manager has not "
                         "confirmed it yet. Checks begin once it is confirmed.",
            "rejected": "The manager sent this list back. It must be corrected, "
                        "re-submitted and confirmed before monitoring starts.",
        }.get(dl.state if dl else "none", "No confirmed list for this date.")
        return jsonify(
            date=day, shift={"id": shift.id, "name": shift.name,
                             "start": shift.start_hhmm, "end": shift.end_hhmm},
            trucks=0, dormant=True, dormant_reason=why, checks=[],
            list_state=(dl.state if dl else "none"), role=_role(),
            subcontractor_id=sub_id,
        )

    # Only trucks actually SENT are monitored. A truck held back as pending was
    # never dispatched, so chasing it through the checkpoints would be nonsense.
    expected = [r.plate for r in DailyListRow.query.filter_by(list_id=dl.id, state="approved")
                .order_by(DailyListRow.plate).all()]
    if not expected:
        return jsonify(
            date=day, shift={"id": shift.id, "name": shift.name,
                             "start": shift.start_hhmm, "end": shift.end_hhmm},
            trucks=0, dormant=True,
            dormant_reason="The confirmed list has no trucks marked ready.",
            checks=[], list_state=dl.state, role=_role(),
        )

    visits, roles = _visits_and_roles()
    deadline = _shift_deadline(day, shift)
    overdue = datetime.utcnow() >= deadline

    # A checkpoint we have never once observed is blind, not failing. Never let a
    # blind checkpoint report trucks as 'missed' - that manufactures alarms.
    seen_anchor = {v["anchor_id"] for v in visits}

    by_plate_visits = {}
    for v in visits:
        by_plate_visits.setdefault(engine.norm_plate(v["plate"]), []).append(v)

    def evidence(plate, role_name, edge):
        aid = roles.get(role_name)
        if not aid:
            return None
        for v in sorted(by_plate_visits.get(engine.norm_plate(plate), []),
                        key=lambda x: x["enter"]):
            if v["anchor_id"] != aid:
                continue
            if v["enter"] < lo or v["enter"] > lo + CYCLE_SPAN:
                continue
            return v.get(edge) or v.get("enter")
        return None

    out_checks = []
    for chk in checks:
        role_name, edge, phrase = CHECK_SPEC.get(chk.code, (None, "enter", chk.code))
        aid = roles.get(role_name) if role_name else None
        blind = (aid is None) or (aid not in seen_anchor)

        done, pending, missed, unverified = [], [], [], []
        for plate in expected:
            when = evidence(plate, role_name, edge) if role_name else None
            if when:
                done.append({"plate": plate, "at": _fmt(when)})
            elif blind:
                unverified.append({"plate": plate})
            elif overdue:
                missed.append({"plate": plate})
            else:
                pending.append({"plate": plate})

        out_checks.append({
            "code": chk.code,
            "label": chk.label or phrase,
            "blind": blind,
            "blind_note": ("This checkpoint has never been observed by GPS. Trucks "
                           "are shown as unverified, not missed - we cannot see "
                           "here, which is not the same as a truck failing to "
                           "arrive." if blind else ""),
            "done": done, "pending": pending, "missed": missed,
            "unverified": unverified,
            "done_count": len(done), "pending_count": len(pending),
            "missed_count": len(missed), "unverified_count": len(unverified),
        })

    return jsonify(
        date=day,
        shift={"id": shift.id, "name": shift.name,
               "start": shift.start_hhmm, "end": shift.end_hhmm},
        trucks=len(expected),
        dormant=False,
        shift_over=overdue,
        deadline=_fmt(deadline),
        checks=out_checks,
        list_state=dl.state,
        confirmed_by=dl.confirmed_by,
        role=_role(),
    )
