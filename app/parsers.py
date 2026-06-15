"""
xlsx parsers. Each takes a file path / stream and returns plain dicts that map
onto the DB models. Column detection mirrors the original JS (header-name based).
"""
import re
from datetime import datetime, timedelta
from openpyxl import load_workbook

from .engine import norm_plate_strict


def _norm_hdr(c):
    return re.sub(r"\s+", " ", str(c if c is not None else "")).strip().lower()


def _to_date(v):
    if v is None or v == "":
        return None
    if isinstance(v, datetime):
        return v
    if isinstance(v, (int, float)):
        # Excel serial → datetime
        try:
            return datetime(1899, 12, 30) + timedelta(days=float(v))
        except Exception:
            return None
    s = str(v).strip().replace(" ", "T")
    for fmt in ("%Y-%m-%dT%H:%M:%S", "%Y-%m-%dT%H:%M", "%Y-%m-%d",
                "%d/%m/%YT%H:%M:%S", "%d/%m/%YT%H:%M", "%d/%m/%Y"):
        try:
            return datetime.strptime(s, fmt)
        except ValueError:
            continue
    return None


def _combine(d_cell, t_cell):
    d = _to_date(d_cell)
    if not d:
        return None
    hh = mm = ss = 0
    if isinstance(t_cell, datetime):
        hh, mm, ss = t_cell.hour, t_cell.minute, t_cell.second
    elif isinstance(t_cell, (int, float)):
        frac = float(t_cell) - int(t_cell)
        tot = round(frac * 86400)
        hh, mm, ss = tot // 3600, (tot % 3600) // 60, tot % 60
    elif isinstance(t_cell, str) and t_cell.strip():
        m = re.search(r"(\d{1,2}):(\d{2})(?::(\d{2}))?", t_cell)
        if m:
            hh, mm, ss = int(m.group(1)), int(m.group(2)), int(m.group(3) or 0)
    return d.replace(hour=hh, minute=mm, second=ss, microsecond=0)


def _rows(ws):
    return [list(r) for r in ws.iter_rows(values_only=True)]


def parse_dispatch_plan(path):
    wb = load_workbook(path, data_only=True)
    ws = wb["DISPATCH PLAN"] if "DISPATCH PLAN" in wb.sheetnames else \
        wb[next((n for n in wb.sheetnames if re.search(r"dispatch", n, re.I)),
                wb.sheetnames[0])]
    rows = _rows(ws)
    hdr = -1
    for i in range(min(len(rows), 10)):
        if any("load start" in _norm_hdr(c) for c in rows[i]):
            hdr = i
            break
    if hdr < 0:
        hdr = 2
    H = [_norm_hdr(c) for c in rows[hdr]]

    def col(rx):
        return next((i for i, c in enumerate(H) if re.search(rx, c)), -1)

    c_plate = col(r"license plate")
    c_load = col(r"load start")
    c_port = col(r"arrive port")
    out = []
    for r in rows[hdr + 1:]:
        raw = r[c_plate] if c_plate >= 0 and c_plate < len(r) else ""
        key = norm_plate_strict(raw)
        if not key:
            continue
        out.append({
            "plate": str(raw).strip(), "key": key,
            "load_start": _to_date(r[c_load]) if c_load >= 0 and c_load < len(r) else None,
            "port_arrive": _to_date(r[c_port]) if c_port >= 0 and c_port < len(r) else None,
        })
    return out


def parse_load_actual(path):
    wb = load_workbook(path, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = _rows(ws)
    if not rows:
        return []
    H = [_norm_hdr(c) for c in rows[0]]

    def col(rx):
        return next((i for i, c in enumerate(H) if re.search(rx, c)), -1)

    c_plate = col(r"truck\s*no|bi\u1ec3n")
    c_din = col(r"date\s*in")
    c_tin = col(r"time\s*in")
    c_net = col(r"net\s*weight")
    c_tk = col(r"ticketid|ticket")
    out = []
    for r in rows[1:]:
        raw = r[c_plate] if 0 <= c_plate < len(r) else ""
        key = norm_plate_strict(raw)
        if not key:
            continue
        net = None
        if 0 <= c_net < len(r):
            s = re.sub(r"[^0-9.\-]", "", str(r[c_net]))
            net = float(s) if s else None
        out.append({
            "plate": str(raw).strip(), "key": key,
            "load_in": _combine(r[c_din] if 0 <= c_din < len(r) else None,
                                r[c_tin] if 0 <= c_tin < len(r) else None),
            "net": net,
            "ticket": str(r[c_tk]).strip() if 0 <= c_tk < len(r) else "",
        })
    return out


def parse_gps(path):
    """
    Generic GPS parser: detects plate / datetime / lat / lng / speed by header name.
    Handles the common vendor export shapes; extend as new formats appear.
    """
    wb = load_workbook(path, data_only=True)
    out = []
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        rows = _rows(ws)
        if not rows:
            continue
        H = [_norm_hdr(c) for c in rows[0]]

        def col(rx):
            return next((i for i, c in enumerate(H) if re.search(rx, c)), -1)

        c_plate = col(r"plate|license|truck|bi\u1ec3n|device|name")
        c_dt = col(r"time|date|gps\s*time|timestamp")
        c_lat = col(r"^lat|latitude")
        c_lng = col(r"^lng|^lon|longitude")
        c_spd = col(r"velocity|speed")
        c_sta = col(r"status|state")
        if min(c_plate, c_dt, c_lat, c_lng) < 0:
            continue
        for r in rows[1:]:
            try:
                plate = str(r[c_plate]).strip()
                dt = _to_date(r[c_dt])
                lat = float(r[c_lat]); lng = float(r[c_lng])
            except (TypeError, ValueError):
                continue
            if not plate or dt is None:
                continue
            spd = 0.0
            if 0 <= c_spd < len(r):
                try:
                    spd = float(r[c_spd])
                except (TypeError, ValueError):
                    spd = 0.0
            out.append({"plate": plate, "dt": dt, "lat": lat, "lng": lng,
                        "speed": spd,
                        "status": str(r[c_sta]).strip() if 0 <= c_sta < len(r) else "",
                        "source": sheet})
    return out


def parse_subfleet(path):
    wb = load_workbook(path, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = _rows(ws)
    if not rows:
        return []
    H = [_norm_hdr(c) for c in rows[0]]

    def col(rx):
        return next((i for i, c in enumerate(H) if re.search(rx, c)), -1)

    c_plate = col(r"plate|license|truck|bi\u1ec3n")
    c_haul = col(r"haul|direction")
    c_arr = col(r"arrive\s*mine|claimed")
    out = []
    for r in rows[1:]:
        raw = r[c_plate] if 0 <= c_plate < len(r) else ""
        key = norm_plate_strict(raw)
        if not key:
            continue
        out.append({
            "plate": str(raw).strip(), "key": key,
            "declared_haul": str(r[c_haul]).strip() if 0 <= c_haul < len(r) else "",
            "claimed_arrive_mine": _to_date(r[c_arr]) if 0 <= c_arr < len(r) else None,
        })
    return out
