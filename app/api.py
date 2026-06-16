"""
JSON API. All endpoints require login. Data is shared across the team.
Uploads parse server-side and replace the relevant table.
"""
import json
from functools import wraps
from datetime import datetime, date
import openpyxl
from flask import Blueprint, request, jsonify, abort
from flask_login import login_required, current_user
from sqlalchemy import func

from .models import (db, User, Truck, GpsPing, Anchor, RouteLeg, KVStore, LoginEvent, AreaTime,
                     DispatchPlanRow, LoadActualRow, SubFleetRow)
from . import parsers, engine

bp = Blueprint("api", __name__, url_prefix="/api")


# ── Uploads ──────────────────────────────────────────────────────────────────
def _save_tmp():
    f = request.files.get("file")
    if not f:
        return None, (jsonify(error="no file"), 400)
    path = "/tmp/" + (f.filename or "upload.xlsx")
    f.save(path)
    return path, None


@bp.post("/upload/gps")
@login_required
def upload_gps():
    path, err = _save_tmp()
    if err:
        return err
    pings = parsers.parse_gps(path)
    if request.args.get("replace") == "1":
        GpsPing.query.delete()
    for p in pings:
        db.session.add(GpsPing(plate=p["plate"], dt=p["dt"], lat=p["lat"],
                               lng=p["lng"], speed=p["speed"],
                               status=p["status"], source=p["source"]))
    db.session.commit()
    return jsonify(added=len(pings))


@bp.post("/upload/plan")
@login_required
def upload_plan():
    path, err = _save_tmp()
    if err:
        return err
    rows = parsers.parse_dispatch_plan(path)
    DispatchPlanRow.query.delete()
    for r in rows:
        db.session.add(DispatchPlanRow(plate=r["plate"], key=r["key"],
                                       load_start=r["load_start"],
                                       port_arrive=r["port_arrive"]))
    db.session.commit()
    return jsonify(added=len(rows))


@bp.post("/upload/load")
@login_required
def upload_load():
    path, err = _save_tmp()
    if err:
        return err
    rows = parsers.parse_load_actual(path)
    LoadActualRow.query.delete()
    for r in rows:
        db.session.add(LoadActualRow(plate=r["plate"], key=r["key"],
                                     load_in=r["load_in"], net=r["net"],
                                     ticket=r["ticket"]))
    db.session.commit()
    return jsonify(added=len(rows))


@bp.post("/upload/subfleet")
@login_required
def upload_subfleet():
    path, err = _save_tmp()
    if err:
        return err
    rows = parsers.parse_subfleet(path)
    SubFleetRow.query.delete()
    for r in rows:
        db.session.add(SubFleetRow(plate=r["plate"], key=r["key"],
                                   declared_haul=r["declared_haul"],
                                   claimed_arrive_mine=r["claimed_arrive_mine"]))
    db.session.commit()
    return jsonify(added=len(rows))


# ── Fleet ────────────────────────────────────────────────────────────────────
@bp.get("/fleet")
@login_required
def fleet():
    trucks = Truck.query.order_by(Truck.plate).all()
    return jsonify([{"plate": t.plate, "status": t.status, "phone": t.phone,
                     "gps_provider": t.gps_provider, "eff_from": t.eff_from,
                     "eff_to": t.eff_to} for t in trucks])


@bp.post("/fleet")
@login_required
def fleet_upsert():
    d = request.get_json(force=True)
    plate = engine.norm_plate(d.get("plate"))
    if not plate:
        return jsonify(error="plate required"), 400
    t = Truck.query.filter_by(plate=plate).first() or Truck(plate=plate)
    t.status = d.get("status", t.status or "online")
    t.phone = d.get("phone", t.phone or "")
    t.gps_provider = d.get("gps_provider", t.gps_provider or "")
    t.eff_from = d.get("eff_from", t.eff_from or "")
    t.eff_to = d.get("eff_to", t.eff_to or "")
    db.session.add(t)
    db.session.commit()
    return jsonify(ok=True)


@bp.delete("/fleet/<plate>")
@login_required
def fleet_delete(plate):
    t = Truck.query.filter_by(plate=engine.norm_plate(plate)).first()
    if t:
        db.session.delete(t)
        db.session.commit()
    return jsonify(ok=True)


# ── Anchors & routes (read for map; write for config) ────────────────────────
@bp.get("/anchors")
@login_required
def anchors():
    return jsonify([{"id": a.id, "name": a.name, "color": a.color,
                     "category": a.category, "role": a.role,
                     "polygon": a.polygon, "min_dwell_min": a.min_dwell_min}
                    for a in Anchor.query.all()])


@bp.get("/routes")
@login_required
def routes():
    return jsonify([{"leg_key": r.leg_key, "label": r.label,
                     "points": r.points, "speed": r.speed}
                    for r in RouteLeg.query.all()])


@bp.put("/routes/<leg_key>")
@login_required
def route_update(leg_key):
    d = request.get_json(force=True)
    r = RouteLeg.query.filter_by(leg_key=leg_key).first()
    if not r:
        return jsonify(error="unknown leg"), 404
    if "points" in d:
        r.points = d["points"]
    if "speed" in d and d["speed"]:
        r.speed = float(d["speed"])
    db.session.commit()
    return jsonify(ok=True)


# ── Status (the engine output) ───────────────────────────────────────────────
def _last_pings():
    """Latest ping per plate."""
    sub = (db.session.query(GpsPing.plate, func.max(GpsPing.dt).label("mx"))
           .group_by(GpsPing.plate).subquery())
    q = (db.session.query(GpsPing)
         .join(sub, (GpsPing.plate == sub.c.plate) & (GpsPing.dt == sub.c.mx)))
    out = {}
    for p in q:
        out[p.plate] = {"lat": p.lat, "lng": p.lng, "dt": p.dt,
                        "speed": p.speed, "status": p.status}
    return out


@bp.get("/status")
@login_required
def status():
    anchors = [{"id": a.id, "name": a.name, "polygon": a.polygon,
                "min_dwell_min": a.min_dwell_min} for a in Anchor.query.all()]
    roles = {a.role: a.id for a in Anchor.query.all() if a.role}
    routes = {r.leg_key: {"points": r.points, "speed": r.speed}
              for r in RouteLeg.query.all()}

    pings = [{"plate": p.plate, "dt": p.dt, "lat": p.lat, "lng": p.lng,
              "speed": p.speed, "status": p.status}
             for p in GpsPing.query.order_by(GpsPing.dt).all()]

    deactivated = {engine.norm_plate(t.plate) for t in
                   Truck.query.filter_by(status="deactivated").all()}

    visits = engine.build_visits(pings, anchors, deactivated)
    sequences = engine.recompute_sequences(visits, roles)

    plates = set(t.plate for t in Truck.query.all())
    plates |= set(p["plate"] for p in pings)
    plates -= {p for p in plates if engine.norm_plate(p) in deactivated}

    last_pings = _last_pings()
    plan = [{"plate": r.plate, "key": r.key, "load_start": r.load_start,
             "port_arrive": r.port_arrive} for r in DispatchPlanRow.query.all()]
    load = [{"plate": r.plate, "key": r.key, "load_in": r.load_in,
             "net": r.net, "ticket": r.ticket} for r in LoadActualRow.query.all()]

    rows = engine.build_status_rows(list(plates), last_pings, sequences, routes,
                                    anchors, roles, plan, load, deactivated)
    return jsonify(rows=rows, count=len(rows))

# ── Visits (for the Timeline page) ───────────────────────────────────────────
@bp.get("/visits")
@login_required
def visits():
    from datetime import datetime, timedelta
    anchors = [{"id": a.id, "name": a.name, "polygon": a.polygon,
                "min_dwell_min": a.min_dwell_min} for a in Anchor.query.all()]
    role_by_anchor = {a.id: a.role for a in Anchor.query.all() if a.role}

    q = GpsPing.query
    fr = request.args.get("from")
    to = request.args.get("to")
    plate = request.args.get("plate")
    if fr:
        q = q.filter(GpsPing.dt >= datetime.strptime(fr, "%Y-%m-%d"))
    if to:
        q = q.filter(GpsPing.dt < datetime.strptime(to, "%Y-%m-%d") + timedelta(days=1))
    if plate:
        q = q.filter(GpsPing.plate == plate)
    pings = [{"plate": p.plate, "dt": p.dt, "lat": p.lat, "lng": p.lng,
              "speed": p.speed, "status": p.status}
             for p in q.order_by(GpsPing.dt).all()]

    deactivated = {engine.norm_plate(t.plate) for t in
                   Truck.query.filter_by(status="deactivated").all()}
    vs = engine.build_visits(pings, anchors, deactivated)

    out, mn, mx = [], None, None
    for v in vs:
        e, x = v["enter"], v.get("exit") or v["enter"]
        if mn is None or e < mn:
            mn = e
        if mx is None or x > mx:
            mx = x
        out.append({
            "plate": v["plate"], "anchor_name": v["anchor_name"],
            "role": role_by_anchor.get(v["anchor_id"], ""),
            "enter": e.isoformat(), "exit": x.isoformat(),
            "dur_min": round((x - e).total_seconds() / 60),
            "open": bool(v.get("open")),
        })
    return jsonify(visits=out,
                   min=mn.isoformat() if mn else None,
                   max=mx.isoformat() if mx else None)

# ── Plan vs Actual ───────────────────────────────────────────────────────────
@bp.get("/pva")
@login_required
def pva():
    anchors = [{"id": a.id, "name": a.name, "polygon": a.polygon,
                "min_dwell_min": a.min_dwell_min} for a in Anchor.query.all()]
    roles = {a.role: a.id for a in Anchor.query.all() if a.role}
    pings = [{"plate": p.plate, "dt": p.dt, "lat": p.lat, "lng": p.lng,
              "speed": p.speed, "status": p.status}
             for p in GpsPing.query.order_by(GpsPing.dt).all()]
    deactivated = {engine.norm_plate(t.plate) for t in
                   Truck.query.filter_by(status="deactivated").all()}
    visits = engine.build_visits(pings, anchors, deactivated)
    seqs = engine.recompute_sequences(visits, roles)
    plan = DispatchPlanRow.query.all()

    THRESH = 60  # minutes; on-time window is +/- 1 hour
    rows, scored, ontime = [], 0, 0
    for p in plan:
        cyc = [c for c in seqs if engine.norm_plate_strict(c["plate"]) == p.key]
        chosen = None
        if cyc and p.load_start:
            same = [c for c in cyc if c["xppl_in"]
                    and c["xppl_in"].date() == p.load_start.date()]
            if same:
                chosen = same[0]
            else:
                chosen = min(cyc, key=lambda c: abs((c["xppl_in"] - p.load_start).total_seconds())
                             if c["xppl_in"] else 9e18)
        elif cyc:
            chosen = sorted(cyc, key=lambda c: c["cycle_date"])[-1]

        actual_load = (chosen.get("loading_in") or chosen.get("xppl_in")) if chosen else None
        actual_port = chosen.get("chan_may_in") if chosen else None

        port_delta, status, code = None, "no actual yet", "na"
        if actual_port and p.port_arrive:
            port_delta = round((actual_port - p.port_arrive).total_seconds() / 60)
            scored += 1
            if port_delta < -THRESH:
                status, code = "early", "planned"
                ontime += 1
            elif port_delta > THRESH:
                status, code = "late", "unassigned"
            else:
                status, code = "on-time", "reloaded"
                ontime += 1
        elif actual_port:
            status, code = "actual only", "fronthaul"

        rows.append({
            "plate": p.plate,
            "plan_load": engine.fmt(p.load_start),
            "actual_load": engine.fmt(actual_load),
            "plan_port": engine.fmt(p.port_arrive),
            "actual_port": engine.fmt(actual_port),
            "port_delta": port_delta,
            "status": status, "code": code,
        })
    pct = round(ontime / scored * 100) if scored else None
    return jsonify(rows=rows, ontime_pct=pct, scored=scored, total=len(rows))

# ── Shared engine state for analytics endpoints ──────────────────────────────
def _engine_state():
    anchors = [{"id": a.id, "name": a.name, "polygon": a.polygon,
                "min_dwell_min": a.min_dwell_min} for a in Anchor.query.all()]
    roles = {a.role: a.id for a in Anchor.query.all() if a.role}
    routes = {r.leg_key: {"points": r.points, "speed": r.speed}
              for r in RouteLeg.query.all()}
    pings = [{"plate": p.plate, "dt": p.dt, "lat": p.lat, "lng": p.lng,
              "speed": p.speed, "status": p.status}
             for p in GpsPing.query.order_by(GpsPing.dt).all()]
    deactivated = {engine.norm_plate(t.plate) for t in
                   Truck.query.filter_by(status="deactivated").all()}
    visits = engine.build_visits(pings, anchors, deactivated)
    seqs = engine.recompute_sequences(visits, roles)
    return {"anchors": anchors, "roles": roles, "routes": routes,
            "pings": pings, "deactivated": deactivated,
            "visits": visits, "seqs": seqs}


def _status_rows(st):
    plates = set(t.plate for t in Truck.query.all()) | set(p["plate"] for p in st["pings"])
    plates -= {p for p in plates if engine.norm_plate(p) in st["deactivated"]}
    plan = [{"plate": r.plate, "key": r.key, "load_start": r.load_start,
             "port_arrive": r.port_arrive} for r in DispatchPlanRow.query.all()]
    load = [{"plate": r.plate, "key": r.key, "load_in": r.load_in,
             "net": r.net, "ticket": r.ticket} for r in LoadActualRow.query.all()]
    return engine.build_status_rows(list(plates), _last_pings(), st["seqs"],
                                    st["routes"], st["anchors"], st["roles"],
                                    plan, load, st["deactivated"])


# ── Subcontractor fleet status ───────────────────────────────────────────────
@bp.get("/subfleet")
@login_required
def subfleet():
    st = _engine_state()
    rows_by_plate = {engine.norm_plate_strict(r["plate"]): r for r in _status_rows(st)}
    out = []
    for r in SubFleetRow.query.all():
        sr = rows_by_plate.get(r.key)
        haul = sr["haul"] if sr else "-"
        est = sr["eta_criteria"] if (sr and sr["haul"] == "Backhaul") else ""
        last_seen = sr["last_seen"] if sr else ""
        if not last_seen:
            gps = "No GPS"
        elif haul == "Backhaul":
            gps = "Returning"
        elif haul == "Completed":
            gps = "At mine"
        else:
            gps = "Not returning"
        delta = None
        if est and r.claimed_arrive_mine:
            from datetime import datetime
            try:
                e = datetime.strptime(est, "%Y-%m-%d %H:%M")
                delta = round((e - r.claimed_arrive_mine).total_seconds() / 60)
            except Exception:
                delta = None
        out.append({
            "plate": r.plate, "declared_haul": r.declared_haul or "",
            "status": haul, "location": (sr["direction"] if sr else ""),
            "claimed_arrive_mine": engine.fmt(r.claimed_arrive_mine),
            "est_arrive_mine": est, "delta": delta, "gps_check": gps,
        })
    return jsonify(rows=out, count=len(out))


# ── Performance dashboard ────────────────────────────────────────────────────
@bp.get("/performance")
@login_required
def performance():
    st = _engine_state()
    seqs = st["seqs"]
    completed = [c for c in seqs if c["chan_may_in"]]
    # cycle time = xppl_in -> xppl_r (next mine arrival)
    durs = []
    for c in seqs:
        if c["xppl_in"] and c["xppl_r"]:
            durs.append((c["xppl_r"] - c["xppl_in"]).total_seconds() / 3600.0)
    avg_cycle = round(sum(durs) / len(durs), 1) if durs else None
    rows = _status_rows(st)
    by_haul = {}
    for r in rows:
        by_haul[r["haul"]] = by_haul.get(r["haul"], 0) + 1
    return jsonify(
        trucks=len(rows),
        cycles=len(seqs),
        completed_trips=len(completed),
        avg_cycle_hours=avg_cycle,
        plan_cycle_hours=48,
        by_haul=by_haul,
    )


# ── Daily performance ────────────────────────────────────────────────────────
@bp.get("/daily")
@login_required
def daily():
    st = _engine_state()
    ports = {}
    for c in st["seqs"]:
        if c["chan_may_in"]:
            d = c["chan_may_in"].date().isoformat()
            ports[d] = ports.get(d, 0) + 1
    loads = {}
    tonnes = {}
    for r in LoadActualRow.query.all():
        if r.load_in:
            d = r.load_in.date().isoformat()
            loads[d] = loads.get(d, 0) + 1
            if r.net:
                tonnes[d] = round(tonnes.get(d, 0) + r.net / 1000.0, 1)
    days = sorted(set(list(ports) + list(loads)))
    series = [{"date": d, "port_arrivals": ports.get(d, 0),
               "loads": loads.get(d, 0), "tonnes": tonnes.get(d, 0)} for d in days]
    return jsonify(series=series)

# ── Data tables for the Data-input pages ─────────────────────────────────────
@bp.get("/sequences")
@login_required
def sequences():
    st = _engine_state()
    def f(x): return engine.fmt(x)
    rows = []
    for c in st["seqs"]:
        rows.append({
            "plate": c["plate"], "cycle_date": f(c["cycle_date"]),
            "xppl_in": f(c["xppl_in"]), "loading_in": f(c["loading_in"]),
            "lalay_out_in": f(c["lalay_out_in"]), "ql49_out_in": f(c["ql49_out_in"]),
            "chan_may_in": f(c["chan_may_in"]), "ql49_back_in": f(c["ql49_back_in"]),
            "lalay_back_in": f(c["lalay_back_in"]), "xppl_r": f(c["xppl_r"]),
            "backhaul_type": c["backhaul_type"] or "",
            "complete": bool(c["chan_may_in"] and c["xppl_r"]),
        })
    return jsonify(rows=rows, count=len(rows))


@bp.get("/gps_summary")
@login_required
def gps_summary():
    from sqlalchemy import func
    q = (db.session.query(GpsPing.plate, func.count(GpsPing.id),
                          func.min(GpsPing.dt), func.max(GpsPing.dt))
         .group_by(GpsPing.plate).order_by(GpsPing.plate))
    rows = [{"plate": p, "pings": n, "first": engine.fmt(mn), "last": engine.fmt(mx)}
            for p, n, mn, mx in q]
    total = sum(r["pings"] for r in rows)
    return jsonify(rows=rows, total_pings=total, plates=len(rows))


@bp.get("/plan_rows")
@login_required
def plan_rows():
    rows = [{"plate": r.plate, "load_start": engine.fmt(r.load_start),
             "port_arrive": engine.fmt(r.port_arrive)}
            for r in DispatchPlanRow.query.order_by(DispatchPlanRow.load_start).all()]
    return jsonify(rows=rows, count=len(rows))


@bp.get("/load_rows")
@login_required
def load_rows():
    rows = [{"plate": r.plate, "load_in": engine.fmt(r.load_in),
             "net": r.net, "ticket": r.ticket}
            for r in LoadActualRow.query.order_by(LoadActualRow.load_in).all()]
    return jsonify(rows=rows, count=len(rows))


@bp.get("/truckstatus")
@login_required
def truckstatus():
    """Pivot: plate (rows) x date (cols). Cell = that day's key events."""
    st = _engine_state()
    # collect events per (plate, date)
    cells = {}
    dates = set()

    def add(plate, dt, label):
        if not dt:
            return
        d = dt.date().isoformat()
        dates.add(d)
        cells.setdefault(plate, {}).setdefault(d, [])
        cells[plate][d].append((dt, label))

    for c in st["seqs"]:
        add(c["plate"], c["loading_in"], "Load")
        add(c["plate"], c["lalay_out_in"], "Border")
        add(c["plate"], c["chan_may_in"], "Port")
        add(c["plate"], c["lalay_back_in"], "Return")
        add(c["plate"], c["xppl_r"], "Mine")

    plates = sorted(cells.keys())
    dates = sorted(dates)
    out = {}
    for plate in plates:
        out[plate] = {}
        for d in dates:
            evs = sorted(cells[plate].get(d, []))
            out[plate][d] = " · ".join(f"{lbl} {dt.strftime('%H:%M')}" for dt, lbl in evs)
    return jsonify(plates=plates, dates=dates, cells=out)


@bp.put("/anchors/<int:aid>")
@login_required
def anchor_update(aid):
    a = db.session.get(Anchor, aid)
    if not a:
        return jsonify(error="not found"), 404
    d = request.get_json(force=True)
    if "role" in d:
        a.role = d["role"] or ""
    if "min_dwell_min" in d and d["min_dwell_min"] is not None:
        a.min_dwell_min = int(d["min_dwell_min"])
    db.session.commit()
    return jsonify(ok=True)


# ── Shared key-value store (backs the full tool's localStorage) ──────────────
@bp.get("/kv")
@login_required
def kv_all():
    return jsonify({r.key: r.value for r in KVStore.query.all()})


@bp.put("/kv/<path:key>")
@login_required
def kv_put(key):
    d = request.get_json(force=True, silent=True) or {}
    val = d.get("value", "")
    row = db.session.get(KVStore, key)
    if row is None:
        row = KVStore(key=key, value=val)
        db.session.add(row)
    else:
        row.value = val
    db.session.commit()
    return jsonify(ok=True)


@bp.delete("/kv/<path:key>")
@login_required
def kv_delete(key):
    row = db.session.get(KVStore, key)
    if row is not None:
        db.session.delete(row)
        db.session.commit()
    return jsonify(ok=True)


# ── Access control: identity + admin-managed tab access ──────────────────────
APP_KEYS = ["tms", "report"]

TAB_KEYS = ["perf","daily","timeline","pva","subfleet","anchors","data","gps","plan",
            "truckstatus","weigh","fleet","guide"]


def _tabs_of(u):
    if not u.allowed_tabs:
        return None  # None = all tabs allowed
    try:
        return [t for t in json.loads(u.allowed_tabs) if t in TAB_KEYS]
    except Exception:
        return None


def _apps_of(u):
    if not getattr(u, "allowed_apps", None):
        return None  # None = all apps allowed
    try:
        return [a for a in json.loads(u.allowed_apps) if a in APP_KEYS]
    except Exception:
        return None


def admin_required(f):
    @wraps(f)
    @login_required
    def w(*a, **k):
        if not getattr(current_user, "is_admin", False):
            abort(403)
        return f(*a, **k)
    return w


@bp.get("/me")
@login_required
def me():
    return jsonify(username=current_user.username,
                   is_admin=bool(current_user.is_admin),
                   tabs=_tabs_of(current_user),
                   lang=(current_user.lang or "en"),
                   default_page=(current_user.default_page or None),
                   can_edit=bool(current_user.can_edit),
                   apps=_apps_of(current_user))


@bp.get("/admin/users")
@admin_required
def admin_users():
    out = []
    for u in User.query.order_by(User.id.asc()).all():
        out.append({"id": u.id, "username": u.username,
                    "is_admin": bool(u.is_admin), "tabs": _tabs_of(u),
                    "lang": (u.lang or "en"), "default_page": (u.default_page or ""),
                    "can_edit": bool(u.can_edit), "apps": _apps_of(u)})
    return jsonify(out)


@bp.put("/admin/users/<int:uid>")
@admin_required
def admin_update(uid):
    u = db.session.get(User, uid)
    if not u:
        abort(404)
    d = request.get_json(force=True, silent=True) or {}
    if "is_admin" in d:
        u.is_admin = bool(d["is_admin"])
    if "tabs" in d:
        t = d["tabs"]
        if t is None:
            u.allowed_tabs = None
        else:
            u.allowed_tabs = json.dumps([str(x) for x in t if str(x) in TAB_KEYS])
    if "lang" in d:
        u.lang = "vi" if str(d["lang"]).lower() == "vi" else "en"
    if "default_page" in d:
        dp = str(d["default_page"] or "")
        u.default_page = dp if dp in TAB_KEYS else None
    if "can_edit" in d:
        u.can_edit = bool(d["can_edit"])
    if "apps" in d:
        a = d["apps"]
        u.allowed_apps = None if a is None else json.dumps([str(x) for x in a if str(x) in APP_KEYS])
    if d.get("password"):
        u.set_password(str(d["password"]))
    db.session.commit()
    # Safety net: never leave the system with zero admins.
    if not User.query.filter_by(is_admin=True).first():
        u.is_admin = True
        db.session.commit()
        return jsonify(ok=True, note="At least one admin is required; kept this user as admin.")
    return jsonify(ok=True, tabs=_tabs_of(u), is_admin=bool(u.is_admin))


@bp.post("/admin/users")
@admin_required
def admin_create():
    d = request.get_json(force=True, silent=True) or {}
    username = (d.get("username") or "").strip()
    password = d.get("password") or ""
    if not username or not password:
        return jsonify(ok=False, error="username and password required"), 400
    if User.query.filter(db.func.lower(User.username) == username.lower()).first():
        return jsonify(ok=False, error="username already exists"), 409
    u = User(username=username)
    u.set_password(str(password))
    u.is_admin = bool(d.get("is_admin", False))
    u.can_edit = bool(d.get("can_edit", True))
    u.lang = "vi" if str(d.get("lang", "en")).lower() == "vi" else "en"
    dp = str(d.get("default_page") or "")
    u.default_page = dp if dp in TAB_KEYS else None
    t = d.get("tabs", None)
    u.allowed_tabs = None if t is None else json.dumps([str(x) for x in t if str(x) in TAB_KEYS])
    ap = d.get("apps", None)
    u.allowed_apps = None if ap is None else json.dumps([str(x) for x in ap if str(x) in APP_KEYS])
    db.session.add(u)
    db.session.commit()
    return jsonify(ok=True, id=u.id)


@bp.delete("/admin/users/<int:uid>")
@admin_required
def admin_delete(uid):
    u = db.session.get(User, uid)
    if not u:
        abort(404)
    if u.id == current_user.id:
        return jsonify(ok=False, error="You cannot delete your own account."), 400
    if u.is_admin and User.query.filter_by(is_admin=True).count() <= 1:
        return jsonify(ok=False, error="Cannot delete the last admin."), 400
    db.session.delete(u)
    db.session.commit()
    return jsonify(ok=True)


# ── Usage analytics ──────────────────────────────────────────────────────────
@bp.post("/track")
@login_required
def track():
    d = request.get_json(force=True, silent=True) or {}
    page = str(d.get("page") or "")[:40]
    try:
        secs = int(d.get("seconds") or 0)
    except Exception:
        secs = 0
    if not page or secs <= 0:
        return jsonify(ok=True)
    if secs > 3600:
        secs = 3600  # guard against absurd gaps
    day = datetime.utcnow().strftime("%Y-%m-%d")
    row = AreaTime.query.filter_by(user_id=current_user.id, page=page, day=day).first()
    if row:
        row.seconds = (row.seconds or 0) + secs
    else:
        db.session.add(AreaTime(user_id=current_user.id, username=current_user.username,
                                page=page, day=day, seconds=secs))
    db.session.commit()
    return jsonify(ok=True)


@bp.get("/admin/activity")
@admin_required
def admin_activity():
    out = {}

    def U(name):
        return out.setdefault(name, {"total_seconds": 0, "areas": {}, "countries": {}, "logins": []})

    for r in AreaTime.query.all():
        u = U(r.username or "?")
        u["areas"][r.page] = u["areas"].get(r.page, 0) + (r.seconds or 0)
        u["total_seconds"] += (r.seconds or 0)

    for e in LoginEvent.query.order_by(LoginEvent.ts.desc()).all():
        u = U(e.username or "?")
        c = e.country or "Unknown"
        cc = u["countries"].setdefault(c, {"code": e.country_code or "", "count": 0})
        cc["count"] += 1
        if len(u["logins"]) < 25:
            u["logins"].append({"ts": (e.ts.isoformat() if e.ts else ""), "ip": e.ip or "", "country": c})

    # shape for the client: sorted areas, country list
    res = {}
    for name, u in out.items():
        areas = sorted(([{"page": p, "seconds": s} for p, s in u["areas"].items()]),
                       key=lambda x: -x["seconds"])
        countries = sorted(([{"country": c, "code": v["code"], "count": v["count"]} for c, v in u["countries"].items()]),
                           key=lambda x: -x["count"])
        res[name] = {"total_seconds": u["total_seconds"], "areas": areas,
                     "countries": countries, "logins": u["logins"]}
    return jsonify(res)


# ── Daily/Weekly Coal Report: contracts + readiness data layer ───────────────
def _kv_get(key):
    row = db.session.get(KVStore, key)
    return row.value if row else None


def _kv_set(key, value):
    row = db.session.get(KVStore, key)
    if row:
        row.value = value
    else:
        db.session.add(KVStore(key=key, value=value))
    db.session.commit()


def _rs(v):
    return "" if v is None else str(v).strip()


def _ymd(v):
    if isinstance(v, (datetime, date)):
        return v.strftime("%Y-%m-%d")
    s = _rs(v).split(" ")[0]
    for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y", "%m/%d/%Y"):
        try:
            return datetime.strptime(s, fmt).strftime("%Y-%m-%d")
        except Exception:
            pass
    return ""


def _hdr_index(headers):
    def norm(x):
        return _rs(x).lower().replace(" ", "").replace(".", "").replace("_", "").replace("#", "")
    return {norm(h): i for i, h in enumerate(headers)}


def _num(v):
    try:
        return float(str(v).replace(",", "") or 0)
    except Exception:
        return 0.0


def editor_required(f):
    @wraps(f)
    @login_required
    def w(*a, **k):
        if not getattr(current_user, "can_edit", False):
            abort(403)
        return f(*a, **k)
    return w


def _open_upload():
    fs = request.files.get("file")
    if not fs:
        abort(400)
    return openpyxl.load_workbook(fs.stream, data_only=True, read_only=True)


@bp.post("/coalrpt/contracts")
@editor_required
def coalrpt_contracts():
    wb = _open_upload()
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return jsonify(ok=False, error="empty file"), 400
    H = _hdr_index(rows[0])
    def g(r, key):
        i = H.get(key)
        return r[i] if i is not None and i < len(r) else None
    out = []
    for r in rows[1:]:
        if not _rs(g(r, "contract")):
            continue
        out.append({"customer": _rs(g(r, "customername")), "export": _rs(g(r, "export")),
                    "term": _rs(g(r, "term")), "endUser": _rs(g(r, "enduser")),
                    "contract": _rs(g(r, "contract")), "start": _ymd(g(r, "start")),
                    "end": _ymd(g(r, "end")), "gar": _rs(g(r, "gartype")),
                    "transporter": _rs(g(r, "transporter")), "qty": _num(g(r, "cntqty")),
                    "tonPerTrip": _num(g(r, "tontrip")), "status": _rs(g(r, "status"))})
    _kv_set("coalRptContracts_v1", json.dumps(out))
    return jsonify(ok=True, count=len(out))


@bp.post("/coalrpt/readiness")
@editor_required
def coalrpt_readiness():
    wb = _open_upload()
    ws = wb["Data"] if "Data" in wb.sheetnames else wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return jsonify(ok=False, error="empty file"), 400
    H = _hdr_index(rows[0])
    def g(r, key):
        i = H.get(key)
        return r[i] if i is not None and i < len(r) else None
    try:
        store = json.loads(_kv_get("coalRptReady_v1") or "{}")
    except Exception:
        store = {}
    added = 0
    for r in rows[1:]:
        rid = _rs(g(r, "id"))
        if not rid:
            continue
        store[rid] = {"id": rid, "regDate": _ymd(g(r, "registerdate")),
                      "contract": _rs(g(r, "contractno")), "plate": _rs(g(r, "truckplate")),
                      "country": _rs(g(r, "country")), "company": _rs(g(r, "transportcompany")),
                      "sub": _rs(g(r, "subcontractor"))}
        added += 1
    _kv_set("coalRptReady_v1", json.dumps(store))
    return jsonify(ok=True, added=added, total=len(store))


@bp.post("/coalrpt/readiness/remove")
@editor_required
def coalrpt_readiness_remove():
    d = request.get_json(force=True, silent=True) or {}
    frm, to = _rs(d.get("from")), _rs(d.get("to"))
    try:
        store = json.loads(_kv_get("coalRptReady_v1") or "{}")
    except Exception:
        store = {}
    before = len(store)
    store = {rid: row for rid, row in store.items()
             if not ((not frm or row.get("regDate", "") >= frm) and (not to or row.get("regDate", "") <= to))}
    _kv_set("coalRptReady_v1", json.dumps(store))
    return jsonify(ok=True, removed=before - len(store), total=len(store))


@bp.get("/coalrpt/data")
@login_required
def coalrpt_data():
    try:
        wmonths = json.loads(_kv_get("coalRptWeigh_v1") or "{}")
    except Exception:
        wmonths = {}
    weigh = []
    for m in wmonths.values():
        if isinstance(m, list):
            weigh.extend(m)
    try:
        contracts = json.loads(_kv_get("coalRptContracts_v1") or "[]")
    except Exception:
        contracts = []
    try:
        ready = list(json.loads(_kv_get("coalRptReady_v1") or "{}").values())
    except Exception:
        ready = []
    return jsonify(weighbridge=weigh, contracts=contracts, readiness=ready)
